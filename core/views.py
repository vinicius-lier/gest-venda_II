# -*- coding: utf-8 -*-
"""
Views do sistema de gestão operacional de vendas
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import AuthenticationForm
from .models import (
    Cliente, Motocicleta, Venda, Consignacao, Seguro, CotacaoSeguro, 
    Seguradora, PlanoSeguro, Bem, Usuario, Loja, Ocorrencia, 
    MenuPerfil, Perfil, MenuUsuario, VendaFinanceira, Despesa, 
    ReceitaExtra, Pagamento, ComunicacaoVenda, Notificacao
)
from .forms import (
    MotocicletaForm, VendaForm, ConsignacaoForm, SeguroForm, 
    CotacaoSeguroForm, SeguradoraForm, PlanoSeguroForm, BemForm, 
    UsuarioForm, LojaForm, OcorrenciaForm, ComentarioOcorrenciaForm, 
    ClienteForm, ComunicacaoVendaForm
)
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.utils import timezone
from datetime import datetime, timedelta
from .importers import DataImporter
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from decimal import Decimal
import logging
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from django.urls import reverse
import os
from django.conf import settings
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import locale

logger = logging.getLogger(__name__)

def login_view(request):
    """View de login do sistema"""
    logger.info(f"Tentativa de login para usuário: {request.POST.get('username', 'N/A')}")
    
    if request.method == 'POST':
        logger.info(f"Formulário válido para usuário: {request.POST.get('username', 'N/A')}")
        
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            
            logger.info(f"Formulário válido para usuário: {username}")
            
            user = authenticate(
                username=username,
                password=password
            )
            
            if user:
                logger.info(f"Usuário autenticado com sucesso: {username}")
                login(request, user)
                logger.info(f"Login realizado para: {username}, redirecionando para dashboard")
                return redirect('core:dashboard')
            else:
                logger.warning(f"Falha na autenticação para usuário: {username}")
                messages.error(request, 'Usuário ou senha inválidos.')
        else:
            logger.warning(f"Formulário inválido para usuário: {request.POST.get('username', 'N/A')} - Erros: {form.errors}")
            messages.error(request, 'Dados inválidos. Verifique usuário e senha.')
    else:
        form = AuthenticationForm()
        logger.debug("Página de login carregada (GET)")
    
    return render(request, 'core/login.html', {'form': form})

def logout_view(request):
    """View de logout do sistema"""
    logout(request)
    messages.success(request, 'Logout realizado com sucesso.')
    return redirect('core:login')

@login_required
def dashboard(request):
    """Dashboard principal do sistema - Otimizado"""
    logger.info(f"Dashboard acessado por: {request.user.username}")
    
    try:
        # Usar cache para estatísticas que não mudam frequentemente
        from django.core.cache import cache
        
        cache_key = f'dashboard_stats_{request.user.id}'
        stats = cache.get(cache_key)
        
        if not stats:
            hoje = timezone.now().date()
            primeiro_dia_mes = hoje.replace(day=1)

            total_clientes = Cliente.objects.filter(ativo=True).count()
            total_motos = Motocicleta.objects.filter(ativo=True).count()
            motos_estoque = Motocicleta.objects.filter(status='estoque', ativo=True).count()
            vendas_mes = Venda.objects.filter(data_venda__gte=primeiro_dia_mes, status='vendido').count()
            consignacoes = Consignacao.objects.filter(status='disponivel', moto__ativo=True).count()

            stats = {
                'total_clientes': total_clientes,
                'total_motos': total_motos,
                'motos_estoque': motos_estoque,
                'vendas_mes': vendas_mes,
                'consignacoes': consignacoes,
            }
            
            # Cache por 5 minutos
            cache.set(cache_key, stats, 300)

        # Dados para gráfico de ranking de vendedores (otimizado)
        ranking_vendedores = Usuario.objects.filter(
            vendas_realizadas__status='vendido'
        ).annotate(
            total_vendas=Count('vendas_realizadas', filter=Q(vendas_realizadas__status='vendido'))
        ).filter(total_vendas__gt=0).order_by('-total_vendas')[:10]

        # Dados para gráfico de motos mais vendidas (otimizado)
        ranking_motos = Motocicleta.objects.filter(
            vendas__status='vendido'
        ).annotate(
            total_vendas=Count('vendas', filter=Q(vendas__status='vendido'))
        ).filter(total_vendas__gt=0).order_by('-total_vendas')[:10]

        # Preparar dados JSON para os gráficos
        ranking_vendedores_labels = json.dumps([
            vendedor.user.first_name or vendedor.user.username 
            for vendedor in ranking_vendedores
        ])
        ranking_vendedores_data = json.dumps([
            vendedor.total_vendas 
            for vendedor in ranking_vendedores
        ])
        ranking_motos_labels = json.dumps([
            f"{moto.marca} {moto.modelo}" 
            for moto in ranking_motos
        ])
        ranking_motos_data = json.dumps([
            moto.total_vendas 
            for moto in ranking_motos
        ])

        context = {
            'ranking_vendedores': ranking_vendedores,
            'ranking_motos': ranking_motos,
            'ranking_vendedores_labels': ranking_vendedores_labels,
            'ranking_vendedores_data': ranking_vendedores_data,
            'ranking_motos_labels': ranking_motos_labels,
            'ranking_motos_data': ranking_motos_data,
            'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
            **stats
        }
        
        logger.info(f"Dashboard carregado com sucesso para: {request.user.username}")
        return render(request, 'core/dashboard.html', context)
        
    except Exception as e:
        logger.error(f"Erro no dashboard para {request.user.username}: {str(e)}")
        messages.error(request, 'Erro ao carregar dashboard. Tente novamente.')
        return render(request, 'core/dashboard.html', {
            'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
        })

@login_required
def cliente_list(request):
    """Lista de clientes com filtros e paginação"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_cliente')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar clientes.'})
    query = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '')
    status = request.GET.get('status', 'ativo')  # Filtro de status
    tipos = Cliente.TIPO_CHOICES
    
    # Filtrar apenas clientes ativos por padrão
    if status == 'ativo':
        clientes = Cliente.objects.filter(ativo=True).order_by('-data_cadastro')
    elif status == 'inativo':
        clientes = Cliente.objects.filter(ativo=False).order_by('-data_cadastro')
    else:
        clientes = Cliente.objects.all().order_by('-data_cadastro')
    
    # Filtro de busca
    if query:
        clientes = clientes.filter(
            Q(nome__icontains=query) |
            Q(cpf_cnpj__icontains=query) |
            Q(telefone__icontains=query)
        )
    # Filtro de tipo
    if tipo:
        clientes = clientes.filter(tipo=tipo)
    
    # Paginação
    paginator = Paginator(clientes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'tipo': tipo,
        'status': status,
        'tipos': tipos,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/cliente_list.html', context)

@login_required
def cliente_create(request):
    """Criação de novo cliente"""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            try:
                # Verificar se o cliente já existe pelo CPF/CNPJ
                cpf_cnpj = form.cleaned_data.get('cpf_cnpj')
                if cpf_cnpj:
                    cliente_existente = Cliente.objects.filter(cpf_cnpj=cpf_cnpj).first()
                    if cliente_existente:
                        if cliente_existente.ativo:
                            messages.error(request, f'Já existe um cliente ativo com este CPF/CNPJ. Caso deseje atualizar os dados, edite o cadastro existente.')
                        else:
                            messages.warning(request, f'Já existe um cliente com este CPF/CNPJ, mas ele está inativo. Você pode reativá-lo abaixo.')
                            # Exibir botão/link para reativar
                            return render(request, 'core/cliente_form.html', {
                                'form': form,
                                'cliente_inativo': cliente_existente,
                                'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
                            })
                        return render(request, 'core/cliente_form.html', {
                            'form': form,
                            'cliente': None,
                            'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
                        })
                cliente = form.save()
                messages.success(request, 'Cliente registrado com sucesso!')
                return redirect('core:cliente_list')
            except Exception as e:
                messages.error(request, f'Erro ao salvar cliente: {str(e)}')
        else:
            # Mostrar erros específicos do formulário
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Erro no campo {field}: {error}')
            if not form.errors:
                messages.error(request, 'Erro ao registrar cliente. Verifique os dados.')
    else:
        form = ClienteForm()
    return render(request, 'core/cliente_form.html', {
        'form': form,
        'cliente': None,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def cliente_detail(request, pk):
    """Detalhes do cliente"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_cliente')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar clientes.'})
    
    cliente = get_object_or_404(Cliente, pk=pk)
    
    # Buscar dados relacionados
    compras = cliente.compras.all().order_by('-data_venda')[:5]
    motos_propriedade = cliente.motos_propriedade.all()
    consignacoes = cliente.consignacoes.all().order_by('-data_entrada')[:5]
    seguros = cliente.seguros.all().order_by('-data_venda')[:5]
    
    context = {
        'cliente': cliente,
        'compras': compras,
        'motos_propriedade': motos_propriedade,
        'consignacoes': consignacoes,
        'seguros': seguros,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/cliente_detail.html', context)

@login_required
def cliente_update(request, pk):
    """Atualização de cliente"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_cliente')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para editar clientes.'})
    
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente atualizado com sucesso!')
            return redirect('core:cliente_detail', pk=cliente.pk)
    else:
        form = ClienteForm(instance=cliente)
    
    return render(request, 'core/cliente_form.html', {
        'form': form,
        'cliente': cliente,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def cliente_delete(request, pk):
    """Exclusão de cliente"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_cliente')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir clientes.'})
    
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        # Verificar se o cliente é proprietário de motocicletas
        motocicletas_proprietario = cliente.motos_propriedade.all()
        
        if motocicletas_proprietario.exists():
            # Registrar no histórico antes de remover
            from .models import HistoricoProprietario
            for moto in motocicletas_proprietario:
                HistoricoProprietario.objects.create(
                    moto=moto,
                    proprietario=cliente,
                    data_inicio=moto.data_entrada,
                    data_fim=timezone.now().date(),
                    motivo='exclusao_cliente',
                    valor_transacao=moto.valor_atual
                )
                # Remover o proprietário da motocicleta
                moto.proprietario = None
                moto.save()
            
            messages.warning(request, f'Proprietário removido de {motocicletas_proprietario.count()} motocicleta(s) antes da exclusão do cliente.')
        
        # Verificar se o cliente é fornecedor de motocicletas
        motocicletas_fornecedor = cliente.motos_fornecidas.all()
        
        if motocicletas_fornecedor.exists():
            # Remover o fornecedor das motocicletas
            for moto in motocicletas_fornecedor:
                moto.fornecedor = None
                moto.save()
            
            messages.warning(request, f'Fornecedor removido de {motocicletas_fornecedor.count()} motocicleta(s) antes da exclusão do cliente.')
        
        # Agora pode excluir o cliente
        cliente.delete()
        messages.success(request, f'Cliente {cliente.nome} excluído com sucesso!')
        return redirect('core:cliente_list')
    
    return render(request, 'core/cliente_confirm_delete.html', {
        'cliente': cliente,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def motocicleta_list(request):
    """Lista de motocicletas com filtros e paginação"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_motocicleta')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar motocicletas.'})
    
    query = request.GET.get('q', '').strip()
    marca = request.GET.get('marca', '')
    modelo = request.GET.get('modelo', '')
    ano = request.GET.get('ano', '')
    status = request.GET.get('status', '')
    ativo = request.GET.get('ativo', 'ativo')  # Filtro de ativo/inativo
    
    # Filtrar apenas motocicletas ativas por padrão
    if ativo == 'ativo':
        motocicletas = Motocicleta.objects.filter(ativo=True).order_by('-data_entrada')
    elif ativo == 'inativo':
        motocicletas = Motocicleta.objects.filter(ativo=False).order_by('-data_entrada')
    else:
        motocicletas = Motocicleta.objects.all().order_by('-data_entrada')
    
    # Filtros
    if query:
        motocicletas = motocicletas.filter(
            Q(marca__icontains=query) |
            Q(modelo__icontains=query) |
            Q(placa__icontains=query) |
            Q(chassi__icontains=query)
        )
    if marca:
        motocicletas = motocicletas.filter(marca__icontains=marca)
    if modelo:
        motocicletas = motocicletas.filter(modelo__icontains=modelo)
    if ano:
        motocicletas = motocicletas.filter(ano=ano)
    if status:
        motocicletas = motocicletas.filter(status=status)
    
    # Paginação
    paginator = Paginator(motocicletas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'motocicletas': page_obj,
        'query': query,
        'marca': marca,
        'modelo': modelo,
        'ano': ano,
        'status': status,
        'ativo': ativo,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/motocicleta_list.html', context)

@login_required
def motocicleta_create(request):
    """Criação de nova motocicleta"""
    if request.method == 'POST':
        form = MotocicletaForm(request.POST, request.FILES)
        if form.is_valid():
            motocicleta = form.save(commit=False)
            motocicleta.save()
            messages.success(request, 'Motocicleta registrada com sucesso!')
            return redirect('core:motocicleta_list')
        else:
            messages.error(request, 'Erro ao registrar motocicleta. Verifique os dados.')
    else:
        form = MotocicletaForm()
    
    return render(request, 'core/motocicleta_form.html', {
        'form': form, 
        'motocicleta': None, 
        'usuario_sistema': request.user
    })

@login_required
def motocicleta_update(request, pk):
    """Atualização de motocicleta"""
    motocicleta = get_object_or_404(Motocicleta, pk=pk)
    if request.method == 'POST':
        form = MotocicletaForm(request.POST, request.FILES, instance=motocicleta)
        if form.is_valid():
            form.save()
            return redirect('core:motocicleta_list')
    else:
        form = MotocicletaForm(instance=motocicleta)
    return render(request, 'core/motocicleta_form.html', {
        'form': form, 
        'motocicleta': motocicleta, 
        'usuario_sistema': request.user
    })

@login_required
def motocicleta_detail(request, pk):
    """Detalhes da motocicleta"""
    motocicleta = get_object_or_404(Motocicleta, pk=pk)
    return render(request, 'core/motocicleta_detail.html', {
        'motocicleta': motocicleta,
        'usuario_sistema': request.user
    })

@login_required
def motocicleta_delete(request, pk):
    """Exclusão de motocicleta"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_motocicleta')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir motocicletas.'})
    
    motocicleta = get_object_or_404(Motocicleta, pk=pk)
    if request.method == 'POST':
        # Verificar se a moto pode ser excluída
        if motocicleta.vendas.exists():
            messages.error(request, 'Não é possível excluir uma motocicleta que possui vendas registradas.')
            return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        
        # Verificar se tem consignação
        try:
            consignacao = motocicleta.consignacao
            if consignacao:
                messages.error(request, 'Não é possível excluir uma motocicleta que possui consignação ativa.')
                return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        except Exception:
            pass  # Não tem consignação, pode continuar
        
        # Verificar se tem seguro
        try:
            seguro = motocicleta.seguro
            if seguro:
                messages.error(request, 'Não é possível excluir uma motocicleta que possui seguro ativo.')
                return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        except Exception:
            pass  # Não tem seguro, pode continuar
        
        if motocicleta.repasses.exists():
            messages.error(request, 'Não é possível excluir uma motocicleta que possui repasses registrados.')
            return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        
        # Verificar se tem controle de chave
        try:
            from administrativo.models import ControleChave
            if ControleChave.objects.filter(motocicleta=motocicleta).exists():
                messages.error(request, 'Não é possível excluir uma motocicleta que possui registros de controle de chave.')
                return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        except ImportError:
            pass  # App administrativo não está disponível, pode continuar
        
        # Marcar como inativo (soft delete)
        motocicleta.ativo = False
        motocicleta.save()
        messages.success(request, 'Motocicleta marcada como inativa com sucesso!')
        return redirect('core:motocicleta_list')
    
    return render(request, 'core/motocicleta_confirm_delete.html', {
        'moto': motocicleta,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def venda_list(request):
    """Lista de vendas com filtros"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_venda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar vendas.'})
    
    # Obter parâmetros de filtro
    status = request.GET.get('status', '')
    vendedor = request.GET.get('vendedor', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    # Query base
    vendas = Venda.objects.all().order_by('-data_venda', '-data_atendimento')
    
    # Aplicar filtros
    if status:
        vendas = vendas.filter(status=status)
    
    if vendedor:
        vendas = vendas.filter(
            Q(vendedor__user__first_name__icontains=vendedor) |
            Q(vendedor__user__last_name__icontains=vendedor) |
            Q(vendedor__user__username__icontains=vendedor)
        )
    
    if data_inicio:
        try:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            vendas = vendas.filter(data_venda__gte=data_inicio)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
            vendas = vendas.filter(data_venda__lte=data_fim)
        except ValueError:
            pass
    
    context = {
        'vendas': vendas,
        'status': status,
        'vendedor': vendedor,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/venda_list.html', context)

@login_required
def venda_create(request):
    """Criação de nova venda"""
    if request.method == 'POST':
        form = VendaForm(request.POST)
        if form.is_valid():
            venda = form.save(commit=False)
            # Definir o vendedor como o usuário logado se não foi especificado
            if not venda.vendedor:
                venda.vendedor = request.user
            
            # Verificar se o status está sendo alterado para 'vendido'
            status_anterior = None
            if venda.pk:
                try:
                    venda_anterior = Venda.objects.get(pk=venda.pk)
                    status_anterior = venda_anterior.status
                except Venda.DoesNotExist:
                    pass
            
            venda.save()
            
            # Tentar associar pré-venda
            pre_venda_associada = venda.associar_pre_venda()
            
            # Se o status foi alterado para 'vendido', criar comunicações obrigatórias
            if venda.status == 'vendido' and status_anterior != 'vendido':
                venda.criar_comunicacoes_obrigatorias(request.user.usuario_sistema)
                if pre_venda_associada:
                    messages.success(request, 'Venda registrada com sucesso! Pré-venda associada e comunicações obrigatórias foram criadas automaticamente.')
                else:
                    messages.success(request, 'Venda registrada com sucesso! Comunicações obrigatórias foram criadas automaticamente.')
            else:
                if pre_venda_associada:
                    messages.success(request, 'Venda registrada com sucesso! Pré-venda associada automaticamente.')
                else:
                    messages.success(request, 'Venda registrada com sucesso!')
            
            # Notificar o setor administrativo sobre a nova venda
            from .models import Usuario, criar_notificacao
            admins = Usuario.objects.filter(
                status='ativo',
                perfil__nome__in=['admin', 'gerente']
            )
            for admin in admins:
                criar_notificacao(
                    usuario=admin,
                    mensagem=f'Nova venda registrada: {venda.moto.marca} {venda.moto.modelo} para {venda.comprador.nome}',
                    link=f'/vendas/{venda.id}/',
                    tipo='venda'
                )
            
            return redirect('core:venda_list')
    else:
        moto_id = request.GET.get('moto')
        initial = {}
        if moto_id:
            try:
                initial['moto'] = int(moto_id)
            except Exception:
                pass
        form = VendaForm(initial=initial)
        # Definir vendedor padrão como usuário logado
        form.fields['vendedor'].initial = request.user
    
    return render(request, 'core/venda_form.html', {
        'form': form,
        'venda': None,
        'usuario_sistema': request.user
    })

@login_required
def venda_update(request, pk):
    """Atualização de venda"""
    venda = get_object_or_404(Venda, pk=pk)
    if request.method == 'POST':
        form = VendaForm(request.POST, instance=venda)
        if form.is_valid():
            # Verificar se o status está sendo alterado para 'vendido'
            status_anterior = venda.status
            venda_atualizada = form.save(commit=False)
            
            # Se o status foi alterado para 'vendido', criar comunicações obrigatórias
            if venda_atualizada.status == 'vendido' and status_anterior != 'vendido':
                venda_atualizada.save()
                venda_atualizada.criar_comunicacoes_obrigatorias(request.user.usuario_sistema)
                messages.success(request, 'Venda atualizada com sucesso! Comunicações obrigatórias foram criadas automaticamente.')
            else:
                form.save()
                # Tentar associar pré-venda se ainda não foi associada
                if not hasattr(venda_atualizada, 'pre_venda_associada'):
                    pre_venda_associada = venda_atualizada.associar_pre_venda()
                    if pre_venda_associada:
                        messages.success(request, 'Venda atualizada com sucesso! Pré-venda associada automaticamente.')
                    else:
                        messages.success(request, 'Venda atualizada com sucesso!')
                else:
                    messages.success(request, 'Venda atualizada com sucesso!')
            
            # Notificar o setor administrativo sobre a atualização da venda
            from .models import Usuario, criar_notificacao
            admins = Usuario.objects.filter(
                status='ativo',
                perfil__nome__in=['admin', 'gerente']
            )
            for admin in admins:
                criar_notificacao(
                    usuario=admin,
                    mensagem=f'Venda atualizada: {venda.moto.marca} {venda.moto.modelo} para {venda.comprador.nome}',
                    link=f'/vendas/{venda.id}/',
                    tipo='venda'
                )
            
            return redirect('core:venda_list')
    else:
        form = VendaForm(instance=venda)
    
    return render(request, 'core/venda_form.html', {
        'form': form,
        'venda': venda,
        'usuario_sistema': request.user
    })

@login_required
def venda_detail(request, pk):
    """Detalhes da venda"""
    venda = get_object_or_404(Venda, pk=pk)
    return render(request, 'core/venda_detail.html', {
        'venda': venda,
        'usuario_sistema': request.user
    })

@login_required
def venda_delete(request, pk):
    """Exclusão de venda"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_venda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir vendas.'})
    
    venda = get_object_or_404(Venda, pk=pk)
    if request.method == 'POST':
        # Verificar se a venda pode ser excluída
        if venda.status == 'vendido':
            messages.error(request, 'Não é possível excluir uma venda já concluída.')
            return redirect('core:venda_detail', pk=venda.pk)
        
        venda.delete()
        messages.success(request, 'Venda excluída com sucesso!')
        return redirect('core:venda_list')
    
    return render(request, 'core/venda_confirm_delete.html', {
        'venda': venda,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def venda_cancel(request, pk):
    """Cancela uma venda, alterando o status para 'cancelado'"""
    venda = get_object_or_404(Venda, pk=pk)
    if not (request.user.is_superuser or request.user.has_perm('core.change_venda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para cancelar vendas.'})

    if venda.status == 'cancelado':
        messages.info(request, 'Esta venda já está cancelada.')
        return redirect('core:venda_detail', pk=pk)

    if request.method == 'POST':
        venda.status = 'cancelado'
        venda.save()
        messages.success(request, 'Venda cancelada com sucesso!')
        return redirect('core:venda_detail', pk=pk)

    return render(request, 'core/venda_confirm_cancel.html', {
        'venda': venda,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def consignacao_list(request):
    """Lista de consignações com filtros"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_consignacao')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar consignações.'})
    
    # Obter parâmetros de filtro
    status = request.GET.get('status', '')
    vendedor = request.GET.get('vendedor', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    # Query base
    consignacoes = Consignacao.objects.all().order_by('-data_entrada')
    
    # Aplicar filtros
    if status:
        consignacoes = consignacoes.filter(status=status)
    
    if vendedor:
        consignacoes = consignacoes.filter(
            Q(vendedor_responsavel__user__first_name__icontains=vendedor) |
            Q(vendedor_responsavel__user__last_name__icontains=vendedor) |
            Q(vendedor_responsavel__user__username__icontains=vendedor)
        )
    
    if data_inicio:
        try:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            consignacoes = consignacoes.filter(data_entrada__gte=data_inicio)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
            consignacoes = consignacoes.filter(data_entrada__lte=data_fim)
        except ValueError:
            pass
    
    context = {
        'consignacoes': consignacoes,
        'status': status,
        'vendedor': vendedor,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/consignacao_list.html', context)

@login_required
def consignacao_create(request):
    """Criação de nova consignação"""
    if request.method == 'POST':
        form = ConsignacaoForm(request.POST)
        if form.is_valid():
            consignacao = form.save(commit=False)
            # Definir o vendedor responsável como o usuário logado se não foi especificado
            if not consignacao.vendedor_responsavel:
                consignacao.vendedor_responsavel = request.user
            consignacao.save()
            messages.success(request, 'Consignação registrada com sucesso!')
            return redirect('core:consignacao_list')
    else:
        form = ConsignacaoForm()
        # Definir vendedor responsável padrão como usuário logado
        form.fields['vendedor_responsavel'].initial = request.user
    
    return render(request, 'core/consignacao_form.html', {
        'form': form,
        'consignacao': None,
        'usuario_sistema': request.user
    })

@login_required
def consignacao_update(request, pk):
    """Atualização de consignação"""
    consignacao = get_object_or_404(Consignacao, pk=pk)
    if request.method == 'POST':
        form = ConsignacaoForm(request.POST, instance=consignacao)
        if form.is_valid():
            form.save()
            messages.success(request, 'Consignação atualizada com sucesso!')
            return redirect('core:consignacao_list')
    else:
        form = ConsignacaoForm(instance=consignacao)
    
    return render(request, 'core/consignacao_form.html', {
        'form': form,
        'consignacao': consignacao,
        'usuario_sistema': request.user
    })

@login_required
def consignacao_detail(request, pk):
    """Detalhes da consignação"""
    consignacao = get_object_or_404(Consignacao, pk=pk)
    return render(request, 'core/consignacao_detail.html', {
        'consignacao': consignacao,
        'usuario_sistema': request.user
    })

@login_required
def consignacao_delete(request, pk):
    """Exclusão de consignação"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_consignacao')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir consignações.'})
    
    consignacao = get_object_or_404(Consignacao, pk=pk)
    if request.method == 'POST':
        # Verificar se a consignação pode ser excluída
        if consignacao.status == 'vendido':
            messages.error(request, 'Não é possível excluir uma consignação já vendida.')
            return redirect('core:consignacao_detail', pk=consignacao.pk)
        
        consignacao.delete()
        messages.success(request, 'Consignação excluída com sucesso!')
        return redirect('core:consignacao_list')
    
    return render(request, 'core/consignacao_confirm_delete.html', {
        'consignacao': consignacao,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def seguro_list(request):
    """Lista de seguros com filtros"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_seguro')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar seguros.'})
    
    # Obter parâmetros de filtro
    status = request.GET.get('status', '')
    vendedor = request.GET.get('vendedor', '')
    seguradora = request.GET.get('seguradora', '')
    data_vencimento = request.GET.get('data_vencimento', '')
    
    # Query base
    seguros = Seguro.objects.all().order_by('-data_venda')
    
    # Aplicar filtros
    if status:
        seguros = seguros.filter(status=status)
    
    if vendedor:
        seguros = seguros.filter(
            Q(vendedor__user__first_name__icontains=vendedor) |
            Q(vendedor__user__last_name__icontains=vendedor) |
            Q(vendedor__user__username__icontains=vendedor)
        )
    
    if seguradora:
        seguros = seguros.filter(
            Q(plano__seguradora__nome__icontains=seguradora)
        )
    
    if data_vencimento:
        try:
            data_vencimento = datetime.strptime(data_vencimento, '%Y-%m-%d').date()
            seguros = seguros.filter(data_fim=data_vencimento)
        except ValueError:
            pass
    
    context = {
        'seguros': seguros,
        'status': status,
        'vendedor': vendedor,
        'seguradora': seguradora,
        'data_vencimento': data_vencimento,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/seguro_list.html', context)

@login_required
def seguro_create(request):
    """Criação de novo seguro"""
    if request.method == 'POST':
        form = SeguroForm(request.POST)
        if form.is_valid():
            seguro = form.save(commit=False)
            # Definir o vendedor como o usuário logado se não foi especificado
            if not seguro.vendedor:
                seguro.vendedor = request.user
            seguro.save()
            messages.success(request, 'Seguro registrado com sucesso!')
            return redirect('core:seguro_list')
    else:
        form = SeguroForm()
        # Definir vendedor padrão como usuário logado
        form.fields['vendedor'].initial = request.user
    
    return render(request, 'core/seguro_form.html', {
        'form': form,
        'seguro': None,
        'usuario_sistema': request.user
    })

@login_required
def seguro_update(request, pk):
    """Atualização de seguro"""
    seguro = get_object_or_404(Seguro, pk=pk)
    if request.method == 'POST':
        form = SeguroForm(request.POST, instance=seguro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Seguro atualizado com sucesso!')
            return redirect('core:seguro_list')
    else:
        form = SeguroForm(instance=seguro)
    
    return render(request, 'core/seguro_form.html', {
        'form': form,
        'seguro': seguro,
        'usuario_sistema': request.user
    })

@login_required
def seguro_detail(request, pk):
    """Detalhes do seguro"""
    seguro = get_object_or_404(Seguro, pk=pk)
    return render(request, 'core/seguro_detail.html', {
        'seguro': seguro,
        'usuario_sistema': request.user
    })

@login_required
def seguro_delete(request, pk):
    """Exclusão de seguro"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_seguro')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir seguros.'})
    
    seguro = get_object_or_404(Seguro, pk=pk)
    if request.method == 'POST':
        # Verificar se o seguro pode ser excluído
        if seguro.status == 'ativo':
            messages.error(request, 'Não é possível excluir um seguro ativo. Cancele-o primeiro.')
            return redirect('core:seguro_detail', pk=seguro.pk)
        
        seguro.delete()
        messages.success(request, 'Seguro excluído com sucesso!')
        return redirect('core:seguro_list')
    
    return render(request, 'core/seguro_confirm_delete.html', {
        'seguro': seguro,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def cotacao_seguro_list(request):
    """Lista de cotações de seguro"""
    cotacoes = CotacaoSeguro.objects.all().order_by('-data_cotacao')
    return render(request, 'core/cotacao_seguro_list.html', {'cotacoes': cotacoes})

@login_required
def cotacao_seguro_create(request):
    """Criação de nova cotação de seguro"""
    if request.method == 'POST':
        form = CotacaoSeguroForm(request.POST)
        if form.is_valid():
            cotacao = form.save(commit=False)
            # Definir o consultor como o usuário logado se não foi especificado
            if not cotacao.consultor:
                cotacao.consultor = request.user
            cotacao.save()
            messages.success(request, 'Cotação de seguro registrada com sucesso!')
            return redirect('core:cotacao_seguro_list')
    else:
        form = CotacaoSeguroForm()
        # Definir consultor padrão como usuário logado
        form.fields['consultor'].initial = request.user
    
    return render(request, 'core/cotacao_seguro_form.html', {
        'form': form,
        'cotacao': None,
        'usuario_sistema': request.user
    })

@login_required
def cotacao_seguro_update(request, pk):
    """Atualização de cotação de seguro"""
    cotacao = get_object_or_404(CotacaoSeguro, pk=pk)
    if request.method == 'POST':
        form = CotacaoSeguroForm(request.POST, instance=cotacao)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cotação de seguro atualizada com sucesso!')
            return redirect('core:cotacao_seguro_list')
    else:
        form = CotacaoSeguroForm(instance=cotacao)
    
    return render(request, 'core/cotacao_seguro_form.html', {
        'form': form,
        'cotacao': cotacao,
        'usuario_sistema': request.user
    })

@login_required
def cotacao_seguro_detail(request, pk):
    """Detalhes da cotação de seguro"""
    cotacao = get_object_or_404(CotacaoSeguro, pk=pk)
    return render(request, 'core/cotacao_seguro_detail.html', {
        'cotacao': cotacao,
        'usuario_sistema': request.user
    })

@login_required
def cotacao_seguro_delete(request, pk):
    """Exclusão de cotação de seguro"""
    cotacao = get_object_or_404(CotacaoSeguro, pk=pk)
    if request.method == 'POST':
        cotacao.delete()
        messages.success(request, 'Cotação de seguro excluída com sucesso!')
        return redirect('core:cotacao_seguro_list')
    return render(request, 'core/cotacao_seguro_confirm_delete.html', {
        'cotacao': cotacao,
        'usuario_sistema': request.user
    })

@login_required
def seguradora_list(request):
    """Lista de seguradoras"""
    seguradoras = Seguradora.objects.all().order_by('nome')
    return render(request, 'core/seguradora_list.html', {'seguradoras': seguradoras})

@login_required
def seguradora_create(request):
    """Criação de nova seguradora"""
    if request.method == 'POST':
        form = SeguradoraForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Seguradora registrada com sucesso!')
            return redirect('core:seguradora_list')
    else:
        form = SeguradoraForm()
    
    return render(request, 'core/seguradora_form.html', {
        'form': form,
        'seguradora': None,
        'usuario_sistema': request.user
    })

@login_required
def seguradora_update(request, pk):
    """Atualização de seguradora"""
    seguradora = get_object_or_404(Seguradora, pk=pk)
    if request.method == 'POST':
        form = SeguradoraForm(request.POST, instance=seguradora)
        if form.is_valid():
            form.save()
            messages.success(request, 'Seguradora atualizada com sucesso!')
            return redirect('core:seguradora_list')
    else:
        form = SeguradoraForm(instance=seguradora)
    
    return render(request, 'core/seguradora_form.html', {
        'form': form,
        'seguradora': seguradora,
        'usuario_sistema': request.user
    })

@login_required
def seguradora_detail(request, pk):
    """Detalhes da seguradora"""
    seguradora = get_object_or_404(Seguradora, pk=pk)
    return render(request, 'core/seguradora_detail.html', {
        'seguradora': seguradora,
        'usuario_sistema': request.user
    })

@login_required
def seguradora_delete(request, pk):
    """Exclusão de seguradora"""
    seguradora = get_object_or_404(Seguradora, pk=pk)
    if request.method == 'POST':
        seguradora.delete()
        messages.success(request, 'Seguradora excluída com sucesso!')
        return redirect('core:seguradora_list')
    return render(request, 'core/seguradora_confirm_delete.html', {
        'seguradora': seguradora,
        'usuario_sistema': request.user
    })

@login_required
def plano_seguro_list(request):
    """Lista de planos de seguro"""
    planos = PlanoSeguro.objects.all().order_by('seguradora__nome', 'nome')
    return render(request, 'core/plano_seguro_list.html', {'planos': planos})

@login_required
def plano_seguro_create(request):
    """Criação de novo plano de seguro"""
    if request.method == 'POST':
        form = PlanoSeguroForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plano de seguro registrado com sucesso!')
            return redirect('core:plano_seguro_list')
    else:
        form = PlanoSeguroForm()
    
    return render(request, 'core/plano_seguro_form.html', {
        'form': form,
        'plano': None,
        'usuario_sistema': request.user
    })

@login_required
def plano_seguro_update(request, pk):
    """Atualização de plano de seguro"""
    plano = get_object_or_404(PlanoSeguro, pk=pk)
    if request.method == 'POST':
        form = PlanoSeguroForm(request.POST, instance=plano)
        if form.is_valid():
            form.save()
            messages.success(request, 'Plano de seguro atualizado com sucesso!')
            return redirect('core:plano_seguro_list')
    else:
        form = PlanoSeguroForm(instance=plano)
    
    return render(request, 'core/plano_seguro_form.html', {
        'form': form,
        'plano': plano,
        'usuario_sistema': request.user
    })

@login_required
def plano_seguro_detail(request, pk):
    """Detalhes do plano de seguro"""
    plano = get_object_or_404(PlanoSeguro, pk=pk)
    return render(request, 'core/plano_seguro_detail.html', {
        'plano': plano,
        'usuario_sistema': request.user
    })

@login_required
def plano_seguro_delete(request, pk):
    """Exclusão de plano de seguro"""
    plano = get_object_or_404(PlanoSeguro, pk=pk)
    if request.method == 'POST':
        plano.delete()
        messages.success(request, 'Plano de seguro excluído com sucesso!')
        return redirect('core:plano_seguro_list')
    return render(request, 'core/plano_seguro_confirm_delete.html', {
        'plano': plano,
        'usuario_sistema': request.user
    })

@login_required
def bem_list(request):
    """Lista de bens"""
    bens = Bem.objects.all().order_by('tipo', 'descricao')
    return render(request, 'core/bem_list.html', {'bens': bens})

@login_required
def bem_create(request):
    """Criação de novo bem"""
    if request.method == 'POST':
        form = BemForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Bem registrado com sucesso!')
            return redirect('core:bem_list')
    else:
        form = BemForm()
    
    return render(request, 'core/bem_form.html', {
        'form': form,
        'bem': None,
        'usuario_sistema': request.user
    })

@login_required
def bem_update(request, pk):
    """Atualização de bem"""
    bem = get_object_or_404(Bem, pk=pk)
    if request.method == 'POST':
        form = BemForm(request.POST, instance=bem)
        if form.is_valid():
            form.save()
            messages.success(request, 'Bem atualizado com sucesso!')
            return redirect('core:bem_list')
    else:
        form = BemForm(instance=bem)
    
    return render(request, 'core/bem_form.html', {
        'form': form,
        'bem': bem,
        'usuario_sistema': request.user
    })

@login_required
def bem_detail(request, pk):
    """Detalhes do bem"""
    bem = get_object_or_404(Bem, pk=pk)
    return render(request, 'core/bem_detail.html', {
        'bem': bem,
        'usuario_sistema': request.user
    })

@login_required
def bem_delete(request, pk):
    """Exclusão de bem"""
    bem = get_object_or_404(Bem, pk=pk)
    if request.method == 'POST':
        bem.delete()
        messages.success(request, 'Bem excluído com sucesso!')
        return redirect('core:bem_list')
    return render(request, 'core/bem_confirm_delete.html', {
        'bem': bem,
        'usuario_sistema': request.user
    })

# Views administrativas
@login_required
def usuario_list(request):
    """Lista de usuários do sistema com filtros"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_usuario')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar usuários.'})
    
    # Obter parâmetros de filtro
    search = request.GET.get('search', '')
    perfil = request.GET.get('perfil', '')
    status = request.GET.get('status', 'todos')  # Agora o padrão é 'todos'
    loja = request.GET.get('loja', '')
    
    # Query base
    usuarios = Usuario.objects.all().order_by('user__first_name', 'user__last_name')
    
    # Aplicar filtros
    if search:
        usuarios = usuarios.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    if perfil:
        usuarios = usuarios.filter(perfil__nome=perfil)
    
    if status == 'ativo':
        usuarios = usuarios.filter(status='ativo')
    elif status == 'inativo':
        usuarios = usuarios.filter(status='inativo')
    elif status == 'bloqueado':
        usuarios = usuarios.filter(status='bloqueado')
    # Se status == 'todos', não aplica filtro
    
    if loja:
        usuarios = usuarios.filter(loja_id=loja)
    
    lojas = Loja.objects.filter(ativo=True).order_by('nome')
    
    context = {
        'usuarios': usuarios,
        'lojas': lojas,
        'status': status,
        'search': search,
        'perfil': perfil,
        'loja': loja,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/usuario_list.html', context)

@login_required
def usuario_create(request):
    """Criação de novo usuário"""
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário registrado com sucesso!')
            return redirect('core:usuario_list')
    else:
        form = UsuarioForm()
    
    return render(request, 'core/usuario_form.html', {
        'form': form,
        'usuario': None,
        'usuario_sistema': request.user
    })

@login_required
def usuario_update(request, pk):
    """Atualização de usuário"""
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário atualizado com sucesso!')
            return redirect('core:usuario_list')
    else:
        form = UsuarioForm(instance=usuario)
    
    return render(request, 'core/usuario_form.html', {
        'form': form,
        'usuario': usuario,
        'usuario_sistema': request.user
    })

@login_required
def usuario_detail(request, pk):
    """Detalhes do usuário"""
    usuario = get_object_or_404(Usuario, pk=pk)
    return render(request, 'core/usuario_detail.html', {
        'usuario': usuario,
        'usuario_sistema': request.user
    })

@login_required
def usuario_delete(request, pk):
    """Exclusão de usuário"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_usuario')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir usuários.'})
    
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        # Verificar se o usuário pode ser excluído
        if usuario.vendas_realizadas.exists():
            messages.error(request, 'Não é possível excluir um usuário que possui vendas registradas.')
            return redirect('core:usuario_detail', pk=usuario.pk)
        
        if usuario.consignacoes_responsavel.exists():
            messages.error(request, 'Não é possível excluir um usuário que possui consignações responsável.')
            return redirect('core:usuario_detail', pk=usuario.pk)
        
        if usuario.seguros_vendidos.exists():
            messages.error(request, 'Não é possível excluir um usuário que possui seguros vendidos.')
            return redirect('core:usuario_detail', pk=usuario.pk)
        
        usuario.status = 'inativo'
        usuario.save()
        messages.success(request, 'Usuário marcado como inativo com sucesso!')
        return redirect('core:usuario_list')
    
    return render(request, 'core/usuario_confirm_delete.html', {
        'usuario': usuario,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def loja_list(request):
    """Lista de lojas com filtros"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_loja')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar lojas.'})
    
    # Obter parâmetros de filtro
    search = request.GET.get('search', '')
    cidade = request.GET.get('cidade', '')
    status = request.GET.get('status', 'ativo')
    
    # Query base
    lojas = Loja.objects.all().order_by('nome')
    
    # Aplicar filtros
    if search:
        lojas = lojas.filter(
            Q(nome__icontains=search) |
            Q(cnpj__icontains=search) |
            Q(cidade__icontains=search)
        )
    
    if cidade:
        lojas = lojas.filter(cidade__icontains=cidade)
    
    if status == 'ativo':
        lojas = lojas.filter(ativo=True)
    elif status == 'inativo':
        lojas = lojas.filter(ativo=False)
    # Se status == 'todos', não aplica filtro
    
    context = {
        'lojas': lojas,
        'status': status,
        'search': search,
        'cidade': cidade,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/loja_list.html', context)

@login_required
def loja_create(request):
    """Criação de nova loja"""
    if request.method == 'POST':
        form = LojaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Loja registrada com sucesso!')
            return redirect('core:loja_list')
    else:
        form = LojaForm()
    # Não existe loja ainda, então não há usuários ativos
    return render(request, 'core/loja_form.html', {
        'form': form,
        'loja': None,
        'usuarios_ativos': 0,
        'usuario_sistema': request.user
    })

@login_required
def loja_update(request, pk):
    """Atualização de loja"""
    loja = get_object_or_404(Loja, pk=pk)
    if request.method == 'POST':
        form = LojaForm(request.POST, instance=loja)
        if form.is_valid():
            form.save()
            messages.success(request, 'Loja atualizada com sucesso!')
            return redirect('core:loja_list')
    else:
        form = LojaForm(instance=loja)
    usuarios_ativos = loja.usuarios.filter(status='ativo').count()
    return render(request, 'core/loja_form.html', {
        'form': form,
        'loja': loja,
        'usuarios_ativos': usuarios_ativos,
        'usuario_sistema': request.user
    })

@login_required
def loja_detail(request, pk):
    """Detalhes da loja"""
    loja = get_object_or_404(Loja, pk=pk)
    usuarios_ativos = loja.usuarios.filter(status='ativo').count()
    return render(request, 'core/loja_detail.html', {
        'loja': loja,
        'usuarios_ativos': usuarios_ativos,
        'usuario_sistema': request.user
    })

@login_required
def loja_delete(request, pk):
    """Exclusão de loja"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_loja')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir lojas.'})
    
    loja = get_object_or_404(Loja, pk=pk)
    if request.method == 'POST':
        # Verificar se a loja pode ser excluída
        if loja.usuarios.exists():
            messages.error(request, 'Não é possível excluir uma loja que possui usuários cadastrados.')
            return redirect('core:loja_detail', pk=loja.pk)
        
        if loja.vendas.exists():
            messages.error(request, 'Não é possível excluir uma loja que possui vendas registradas.')
            return redirect('core:loja_detail', pk=loja.pk)
        
        loja.ativo = False
        loja.save()
        messages.success(request, 'Loja marcada como inativa com sucesso!')
        return redirect('core:loja_list')
    
    return render(request, 'core/loja_confirm_delete.html', {
        'loja': loja,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

# Views de ocorrências
@login_required
def ocorrencia_list(request):
    """Lista de ocorrências com filtros e dashboard"""
    from django.utils import timezone
    from datetime import timedelta
    if not (request.user.is_superuser or request.user.has_perm('core.view_ocorrencia')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar ocorrências.'})
    
    # Filtros
    q = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')
    prioridade = request.GET.get('prioridade', '')
    status = request.GET.get('status', '')

    ocorrencias = Ocorrencia.objects.all().order_by('-data_abertura')
    if q:
        ocorrencias = ocorrencias.filter(Q(titulo__icontains=q) | Q(descricao__icontains=q))
    if tipo:
        ocorrencias = ocorrencias.filter(tipo=tipo)
    if prioridade:
        ocorrencias = ocorrencias.filter(prioridade=prioridade)
    if status:
        ocorrencias = ocorrencias.filter(status=status)

    # Cards de resumo
    total_ocorrencias = Ocorrencia.objects.count()
    ocorrencias_ultimos_30 = Ocorrencia.objects.filter(data_abertura__gte=timezone.now()-timedelta(days=30)).count()
    ocorrencias_atrasadas = Ocorrencia.objects.filter(data_limite__lt=timezone.now(), status__in=['aberta', 'em_andamento']).count()
    ocorrencias_criticas = Ocorrencia.objects.filter(prioridade='critica').count()

    # Top solicitantes
    top_solicitantes_qs = (
        Ocorrencia.objects.values('solicitante__user__first_name', 'solicitante__user__last_name', 'loja__nome')
        .annotate(qtd=Count('id'))
        .order_by('-qtd')[:5]
    )
    top_solicitantes = [
        {'nome': f"{s['solicitante__user__first_name']} {s['solicitante__user__last_name']}", 'loja': s['loja__nome'], 'qtd': s['qtd']}
        for s in top_solicitantes_qs
    ]

    # Top responsáveis
    top_responsaveis_qs = (
        Ocorrencia.objects.filter(status='resolvida')
        .values('responsavel__user__first_name', 'responsavel__user__last_name', 'loja__nome')
        .annotate(qtd=Count('id'))
        .order_by('-qtd')[:5]
    )
    top_responsaveis = [
        {'nome': f"{r['responsavel__user__first_name']} {r['responsavel__user__last_name']}", 'loja': r['loja__nome'], 'qtd': r['qtd']}
        for r in top_responsaveis_qs
    ]

    # Agrupamentos
    agrupado_status = {dict(Ocorrencia.STATUS_CHOICES).get(s['status'], s['status']): s['qtd'] for s in Ocorrencia.objects.values('status').annotate(qtd=Count('id'))}
    agrupado_tipo = {dict(Ocorrencia.TIPO_CHOICES).get(s['tipo'], s['tipo']): s['qtd'] for s in Ocorrencia.objects.values('tipo').annotate(qtd=Count('id'))}
    agrupado_loja = {s['loja__nome']: s['qtd'] for s in Ocorrencia.objects.values('loja__nome').annotate(qtd=Count('id'))}

    # Choices para filtros
    tipos = dict(Ocorrencia.TIPO_CHOICES)
    prioridades = dict(Ocorrencia.PRIORIDADE_CHOICES)
    status_list = dict(Ocorrencia.STATUS_CHOICES)

    context = {
        'ocorrencias': ocorrencias,
        'total_ocorrencias': total_ocorrencias,
        'ocorrencias_ultimos_30': ocorrencias_ultimos_30,
        'ocorrencias_atrasadas': ocorrencias_atrasadas,
        'ocorrencias_criticas': ocorrencias_criticas,
        'top_solicitantes': top_solicitantes,
        'top_responsaveis': top_responsaveis,
        'agrupado_status': agrupado_status,
        'agrupado_tipo': agrupado_tipo,
        'agrupado_loja': agrupado_loja,
        'tipos': tipos,
        'prioridades': prioridades,
        'status_list': status_list,
        'request': request,
    }
    return render(request, 'core/ocorrencia_list.html', context)

@login_required
def ocorrencia_create(request):
    """Criação de nova ocorrência"""
    logger.info(f"ocorrencia_create: Usuário {request.user.username} tentando criar ocorrência")
    
    # Verificar se o usuário tem usuario_sistema
    if not hasattr(request.user, 'usuario_sistema') or request.user.usuario_sistema is None:
        logger.error(f"ocorrencia_create: Usuário {request.user.username} não tem usuario_sistema")
        messages.error(request, 'Seu usuário não está vinculado a um perfil do sistema. Contate o administrador.')
        return redirect('core:dashboard')
    
    # Verificar se o usuario_sistema tem loja
    if not hasattr(request.user.usuario_sistema, 'loja') or request.user.usuario_sistema.loja is None:
        logger.error(f"ocorrencia_create: Usuário {request.user.username} não tem loja associada")
        messages.error(request, 'Seu usuário não está vinculado a uma loja. Contate o administrador.')
        return redirect('core:dashboard')
    
    logger.info(f"ocorrencia_create: Usuário {request.user.username} tem usuario_sistema e loja válidos")
    
    if request.method == 'POST':
        form = OcorrenciaForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                ocorrencia = form.save(commit=False)
                ocorrencia.solicitante = request.user.usuario_sistema
                ocorrencia.save()
                
                # Detectar menções (@usuario) na descrição e observações
                from .models import Usuario, criar_notificacao
                import re
                
                texto_completo = f"{ocorrencia.descricao} {ocorrencia.observacoes or ''}"
                mencoes = re.findall(r'@(\w+)', texto_completo)
                
                for username in mencoes:
                    try:
                        usuario_mencoes = Usuario.objects.get(
                            user__username=username,
                            status='ativo'
                        )
                        criar_notificacao(
                            usuario=usuario_mencoes,
                            mensagem=f'Você foi mencionado na ocorrência "{ocorrencia.titulo}" por {request.user.get_full_name()}',
                            link=f'/ocorrencias/{ocorrencia.id}/',
                            tipo='menção'
                        )
                    except Usuario.DoesNotExist:
                        pass
                
                logger.info(f"ocorrencia_create: Ocorrência criada com sucesso por {request.user.username}")
                messages.success(request, 'Ocorrência registrada com sucesso!')
                return redirect('core:ocorrencia_list')
            except Exception as e:
                logger.error(f"ocorrencia_create: Erro ao salvar ocorrência: {str(e)}")
                messages.error(request, f'Erro ao registrar ocorrência: {str(e)}')
        else:
            logger.warning(f"ocorrencia_create: Formulário inválido para {request.user.username}: {form.errors}")
    else:
        try:
            loja_inicial = request.user.usuario_sistema.loja
            form = OcorrenciaForm(initial={'loja': loja_inicial} if loja_inicial else {})
            logger.info(f"ocorrencia_create: Formulário criado para {request.user.username} com loja {loja_inicial}")
        except Exception as e:
            logger.error(f"ocorrencia_create: Erro ao criar formulário: {str(e)}")
            messages.error(request, f'Erro ao carregar formulário: {str(e)}')
            return redirect('core:dashboard')
    
    return render(request, 'core/ocorrencia_form.html', {
        'form': form,
        'ocorrencia': None,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None)
    })

@login_required
def ocorrencia_update(request, pk):
    """Atualização de ocorrência"""
    ocorrencia = get_object_or_404(Ocorrencia, pk=pk)
    if request.method == 'POST':
        form = OcorrenciaForm(request.POST, request.FILES, instance=ocorrencia)
        if form.is_valid():
            ocorrencia_atualizada = form.save(commit=False)
            
            # Detectar menções (@usuario) na descrição e observações
            from .models import Usuario, criar_notificacao
            import re
            
            texto_completo = f"{ocorrencia_atualizada.descricao} {ocorrencia_atualizada.observacoes or ''}"
            mencoes = re.findall(r'@(\w+)', texto_completo)
            
            for username in mencoes:
                try:
                    usuario_mencoes = Usuario.objects.get(
                        user__username=username,
                        status='ativo'
                    )
                    criar_notificacao(
                        usuario=usuario_mencoes,
                        mensagem=f'Você foi mencionado na ocorrência "{ocorrencia_atualizada.titulo}" por {request.user.get_full_name()}',
                        link=f'/ocorrencias/{ocorrencia_atualizada.id}/',
                        tipo='menção'
                    )
                except Usuario.DoesNotExist:
                    pass
            
            form.save()
            messages.success(request, 'Ocorrência atualizada com sucesso!')
            return redirect('core:ocorrencia_list')
    else:
        form = OcorrenciaForm(instance=ocorrencia)
    
    return render(request, 'core/ocorrencia_form.html', {
        'form': form,
        'ocorrencia': ocorrencia,
        'usuario_sistema': request.user
    })

@login_required
def ocorrencia_detail(request, pk):
    """Detalhes da ocorrência"""
    ocorrencia = get_object_or_404(Ocorrencia, pk=pk)
    comentarios = ocorrencia.comentarios.all()
    
    if request.method == 'POST':
        comentario_form = ComentarioOcorrenciaForm(request.POST)
        if comentario_form.is_valid():
            comentario = comentario_form.save(commit=False)
            comentario.ocorrencia = ocorrencia
            comentario.autor = request.user
            comentario.save()
            messages.success(request, 'Comentário adicionado com sucesso!')
            return redirect('core:ocorrencia_detail', pk=pk)
    else:
        comentario_form = ComentarioOcorrenciaForm()
    
    return render(request, 'core/ocorrencia_detail.html', {
        'ocorrencia': ocorrencia,
        'comentarios': comentarios,
        'comentario_form': comentario_form,
        'usuario_sistema': request.user
    })

@login_required
def ocorrencia_delete(request, pk):
    """Excluir ocorrência"""
    ocorrencia = get_object_or_404(Ocorrencia, pk=pk)
    
    if request.method == 'POST':
        try:
            ocorrencia.delete()
            messages.success(request, 'Ocorrência excluída com sucesso!')
            return redirect('core:ocorrencia_list')
        except Exception as e:
            messages.error(request, f'Erro ao excluir ocorrência: {str(e)}')
            return redirect('core:ocorrencia_detail', pk=pk)
    
    return render(request, 'core/ocorrencia_confirm_delete.html', {
        'ocorrencia': ocorrencia,
        'usuario_sistema': request.user
    })

# Views de Importação
@login_required
def import_data(request):
    """Redireciona para a tela de pré-visualização de motocicletas"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')
    
    # Redirecionar diretamente para a tela de pré-visualização de motocicletas
    return redirect('core:preview_import_motocicletas')

@login_required
def import_lojas(request):
    """Importação de lojas"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')
    
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            importer = DataImporter()
            success = importer.import_lojas(request.FILES['file'])
            summary = importer.get_import_summary()
            
            if success:
                messages.success(request, f'Importação concluída! {summary["success_count"]} lojas importadas com sucesso.')
            else:
                messages.warning(request, f'Importação concluída com erros. {summary["success_count"]} lojas importadas, {summary["error_count"]} erros.')
                for error in summary.get('errors', [])[:5]:  # Mostra apenas os primeiros 5 erros
                    messages.error(request, f'Erro: {error}')
            
            return redirect('core:preview_import_motocicletas')
        except Exception as e:
            messages.error(request, f'Erro durante a importação: {str(e)}')
            return redirect('core:preview_import_motocicletas')
    
    return render(request, 'core/import_lojas.html', {
        'usuario_sistema': request.user
    })

@login_required
def import_clientes(request):
    """Importação de clientes"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        try:
            importer = DataImporter()
            
            # Verificar se há mapeamento de colunas (vindo da pré-visualização)
            column_mapping = {}
            temp_file_path = None
            if 'map_nome' in request.POST:
                # Mapeamento personalizado foi enviado - usar arquivo temporário
                if 'temp_file_path_clientes' not in request.session:
                    messages.error(request, 'Arquivo não encontrado. Faça o upload novamente.')
                    return redirect('core:preview_import_clientes')
                
                # Usar o arquivo temporário
                temp_file_path = request.session['temp_file_path_clientes']
                
                # Criar mapeamento de colunas
                column_mapping = {
                    'nome': [request.POST.get('map_nome')],
                    'cpf_cnpj': [request.POST.get('map_cpf_cnpj')],
                    'rg': [request.POST.get('map_rg')],
                    'data_nascimento': [request.POST.get('map_data_nascimento')],
                    'telefone': [request.POST.get('map_telefone')],
                    'email': [request.POST.get('map_email')],
                    'endereco': [request.POST.get('map_endereco')],
                    'cidade': [request.POST.get('map_cidade')],
                    'estado': [request.POST.get('map_estado')],
                    'cep': [request.POST.get('map_cep')],
                    'tipo': [request.POST.get('map_tipo')],
                    'observacoes': [request.POST.get('map_observacoes')]
                }
                
                # Remover campos vazios
                column_mapping = {k: v for k, v in column_mapping.items() if v[0]}
                
                # Limpar sessão
                del request.session['temp_file_path_clientes']
                
            else:
                # Upload direto (sem mapeamento)
                if not request.FILES.get('file'):
                    messages.error(request, 'Nenhum arquivo foi enviado.')
                    return redirect('core:preview_import_clientes')
                temp_file_path = request.FILES['file']
            
            success = importer.import_clientes(temp_file_path, column_mapping)
            summary = importer.get_import_summary()
            
            # Limpar arquivo temporário se existir
            if temp_file_path and isinstance(temp_file_path, str) and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass  # Ignorar erros de limpeza
            
            if success:
                messages.success(request, f'Importação concluída! {summary["success_count"]} clientes importados com sucesso.')
            else:
                messages.warning(request, f'Importação concluída com erros. {summary["success_count"]} clientes importados, {summary["error_count"]} erros.')
                for error in summary.get('errors', [])[:5]:
                    messages.error(request, f'Erro: {error}')
            
            return redirect('core:preview_import_clientes')
        except Exception as e:
            messages.error(request, f'Erro durante a importação: {str(e)}')
            return redirect('core:preview_import_clientes')
    
    return render(request, 'core/import_motocicletas.html', {
        'usuario_sistema': request.user
    })

@login_required
def import_vendas(request):
    """Importação de vendas"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        try:
            importer = DataImporter()
            
            # Verificar se há mapeamento de colunas (vindo da pré-visualização)
            column_mapping = {}
            temp_file_path = None
            if 'map_moto_chassi' in request.POST:
                # Mapeamento personalizado foi enviado - usar arquivo temporário
                if 'temp_file_path_vendas' not in request.session:
                    messages.error(request, 'Arquivo não encontrado. Faça o upload novamente.')
                    return redirect('core:preview_import_vendas')
                
                # Usar o arquivo temporário
                temp_file_path = request.session['temp_file_path_vendas']
                
                # Criar mapeamento de colunas
                column_mapping = {
                    'moto_chassi': [request.POST.get('map_moto_chassi')],
                    'comprador_cpf': [request.POST.get('map_comprador_cpf')],
                    'vendedor_username': [request.POST.get('map_vendedor_username')],
                    'loja_nome': [request.POST.get('map_loja_nome')],
                    'origem': [request.POST.get('map_origem')],
                    'forma_pagamento': [request.POST.get('map_forma_pagamento')],
                    'status': [request.POST.get('map_status')],
                    'valor_venda': [request.POST.get('map_valor_venda')],
                    'valor_entrada': [request.POST.get('map_valor_entrada')],
                    'comissao_vendedor': [request.POST.get('map_comissao_vendedor')],
                    'data_atendimento': [request.POST.get('map_data_atendimento')],
                    'data_venda': [request.POST.get('map_data_venda')],
                    'observacoes': [request.POST.get('map_observacoes')]
                }
                
                # Remover campos vazios
                column_mapping = {k: v for k, v in column_mapping.items() if v[0]}
                
                # Limpar sessão
                del request.session['temp_file_path_vendas']
                
            else:
                # Upload direto (sem mapeamento)
                if not request.FILES.get('file'):
                    messages.error(request, 'Nenhum arquivo foi enviado.')
                    return redirect('core:preview_import_vendas')
                temp_file_path = request.FILES['file']
            
            success = importer.import_vendas(temp_file_path, column_mapping)
            summary = importer.get_import_summary()
            
            # Limpar arquivo temporário se existir
            if temp_file_path and isinstance(temp_file_path, str) and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass  # Ignorar erros de limpeza
            
            if success:
                messages.success(request, f'Importação concluída! {summary["success_count"]} vendas importadas com sucesso.')
            else:
                messages.warning(request, f'Importação concluída com erros. {summary["success_count"]} vendas importadas, {summary["error_count"]} erros.')
                for error in summary.get('errors', [])[:5]:
                    messages.error(request, f'Erro: {error}')
            
            return redirect('core:preview_import_vendas')
        except Exception as e:
            messages.error(request, f'Erro durante a importação: {str(e)}')
            return redirect('core:preview_import_vendas')
    
    return render(request, 'core/import_vendas.html', {
        'usuario_sistema': request.user
    })

@login_required
def import_seguradoras(request):
    """Importação de seguradoras"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')
    
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            importer = DataImporter()
            success = importer.import_seguradoras(request.FILES['file'])
            summary = importer.get_import_summary()
            
            if success:
                messages.success(request, f'Importação concluída! {summary["success_count"]} seguradoras importadas com sucesso.')
            else:
                messages.warning(request, f'Importação concluída com erros. {summary["success_count"]} seguradoras importadas, {summary["error_count"]} erros.')
                for error in summary.get('errors', [])[:5]:
                    messages.error(request, f'Erro: {error}')
            
            return redirect('core:preview_import_motocicletas')
        except Exception as e:
            messages.error(request, f'Erro durante a importação: {str(e)}')
            return redirect('core:preview_import_motocicletas')
    
    return render(request, 'core/import_seguradoras.html', {
        'usuario_sistema': request.user
    })

@login_required
def import_planos_seguro(request):
    """Importação de planos de seguro"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')
    
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            importer = DataImporter()
            success = importer.import_planos_seguro(request.FILES['file'])
            summary = importer.get_import_summary()
            
            if success:
                messages.success(request, f'Importação concluída! {summary["success_count"]} planos importados com sucesso.')
            else:
                messages.warning(request, f'Importação concluída com erros. {summary["success_count"]} planos importados, {summary["error_count"]} erros.')
                for error in summary.get('errors', [])[:5]:
                    messages.error(request, f'Erro: {error}')
            
            return redirect('core:preview_import_motocicletas')
        except Exception as e:
            messages.error(request, f'Erro durante a importação: {str(e)}')
            return redirect('core:preview_import_motocicletas')
    
    return render(request, 'core/import_planos_seguro.html', {
        'usuario_sistema': request.user
    })

@require_GET
def buscar_motocicleta(request):
    id_moto = request.GET.get('id_moto')
    placa = request.GET.get('placa')
    chassi = request.GET.get('chassi')
    moto = None
    if id_moto:
        moto = Motocicleta.objects.filter(id=id_moto, ativo=True).first()
    elif placa:
        moto = Motocicleta.objects.filter(placa__iexact=placa, ativo=True).first()
    elif chassi:
        moto = Motocicleta.objects.filter(chassi__iexact=chassi, ativo=True).first()
    if moto:
        return JsonResponse({
            'id': moto.id,
            'placa': moto.placa,
            'chassi': moto.chassi,
            'descricao': f"ID: {moto.id} | Placa: {moto.placa or '-'} | Chassi: {moto.chassi} | {moto.marca} {moto.modelo} {moto.ano}"
        })
    return JsonResponse({}, status=404)

@login_required
def download_modelo_csv(request, tipo):
    """Download de modelo CSV para importação"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem baixar modelos.')
        return redirect('core:dashboard')
    
    modelos = {
        'clientes': 'clientes_modelo.csv',
        'lojas': 'lojas_modelo.csv',
        'motocicletas': 'motocicletas_modelo.csv',
        'vendas': 'vendas_modelo.csv',
        'seguradoras': 'seguradoras_modelo.csv',
        'planos_seguro': 'planos_seguro_modelo.csv',
    }
    
    if tipo not in modelos:
        messages.error(request, 'Tipo de modelo inválido.')
        return redirect('core:preview_import_motocicletas')
    
    # Primeiro, tentar no diretório modelos_importacao (onde os arquivos realmente estão)
    arquivo_path = os.path.join(settings.BASE_DIR, 'modelos_importacao', modelos[tipo])
    
    if not os.path.exists(arquivo_path):
        # Se não existir, tentar no STATIC_ROOT
        arquivo_path = os.path.join(settings.STATIC_ROOT, 'modelos', modelos[tipo])
    
    if not os.path.exists(arquivo_path):
        # Se ainda não existir, tentar no diretório static do projeto
        arquivo_path = os.path.join(settings.BASE_DIR, 'static', 'modelos', modelos[tipo])
    
    if not os.path.exists(arquivo_path):
        messages.error(request, 'Arquivo de modelo não encontrado.')
        return redirect('core:preview_import_motocicletas')
    
    try:
        with open(arquivo_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{modelos[tipo]}"'
            return response
    except Exception as e:
        logger.error(f"Erro ao ler arquivo de modelo {arquivo_path}: {str(e)}")
        messages.error(request, 'Erro ao ler arquivo de modelo.')
        return redirect('core:preview_import_motocicletas')

@csrf_exempt
@login_required
def preview_import_motocicletas(request):
    """Pré-visualização do CSV para importação de motocicletas"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Salvar o arquivo temporariamente
        import tempfile
        import os
        
        # Usar configuração do settings
        temp_dir = getattr(settings, 'TEMP_DIR', '/tmp')
        
        # Criar diretório se não existir (apenas em desenvolvimento)
        if not os.path.exists(temp_dir):
            try:
                os.makedirs(temp_dir, exist_ok=True)
            except Exception as e:
                logger.error(f"Erro ao criar diretório temporário: {str(e)}")
                messages.error(request, 'Erro ao criar diretório temporário.')
                return redirect('core:preview_import_motocicletas')
        
        temp_file_path = os.path.join(temp_dir, f"temp_{request.user.id}_{file.name}")
        
        try:
            with open(temp_file_path, 'wb+') as temp_file:
                for chunk in file.chunks():
                    temp_file.write(chunk)
            
            # Armazenar o caminho do arquivo temporário na sessão
            request.session['temp_file_path'] = temp_file_path
            
            # Tentar ler o arquivo com diferentes encodings
            encodings = ['latin1', 'utf-8-sig', 'utf-8', 'cp1252', 'iso-8859-1', 'windows-1252']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(temp_file_path, encoding=encoding, nrows=10)
                    break
                except Exception as e:
                    logger.warning(f"Falha ao ler com encoding {encoding}: {str(e)}")
                    continue
            
            if df is None:
                messages.error(request, 'Não foi possível ler o arquivo CSV. Tente salvar como CSV (separado por vírgula) e tente novamente.')
                # Limpar arquivo temporário
                try:
                    os.remove(temp_file_path)
                    del request.session['temp_file_path']
                except:
                    pass
                return redirect('core:preview_import_motocicletas')
            
            colunas = list(df.columns)
            preview = df.head(5).values.tolist()
            
            # Lista de campos para mapeamento
            campos = ['marca', 'modelo', 'ano', 'ano_fabricacao', 'cor', 'placa', 'chassi', 'rodagem', 'valor_entrada', 'valor_atual', 'status', 'observacoes']
            
            return render(request, 'core/preview_import_motocicletas.html', {
                'colunas': colunas,
                'preview': preview,
                'arquivo_nome': file.name,
                'campos': campos,
            })
            
        except Exception as e:
            logger.error(f"Erro ao processar arquivo: {str(e)}")
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            # Limpar arquivo temporário
            try:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                if 'temp_file_path' in request.session:
                    del request.session['temp_file_path']
            except:
                pass
            return redirect('core:preview_import_motocicletas')
    
    # Lista de campos para mapeamento (mesmo quando não há arquivo)
    campos = ['marca', 'modelo', 'ano', 'ano_fabricacao', 'cor', 'placa', 'chassi', 'rodagem', 'valor_entrada', 'valor_atual', 'status', 'observacoes']
    
    # Verificar se há relatório de importação na sessão
    import_report = request.session.get('import_report')
    if import_report:
        # Remover o relatório da sessão após exibir
        del request.session['import_report']
        return render(request, 'core/preview_import_motocicletas.html', {
            'campos': campos,
            'import_report': import_report
        })
    
    return render(request, 'core/preview_import_motocicletas.html', {'campos': campos})

@csrf_exempt
@login_required
def preview_import_clientes(request):
    """Pré-visualização do CSV para importação de clientes"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Salvar o arquivo temporariamente
        import tempfile
        import os
        
        # Criar arquivo temporário
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_uploads')
        os.makedirs(temp_dir, exist_ok=True)
        
        temp_file_path = os.path.join(temp_dir, f"temp_clientes_{request.user.id}_{file.name}")
        with open(temp_file_path, 'wb+') as temp_file:
            for chunk in file.chunks():
                temp_file.write(chunk)
        
        # Armazenar o caminho do arquivo temporário na sessão
        request.session['temp_file_path_clientes'] = temp_file_path
        
        encodings = ['latin1', 'utf-8-sig', 'utf-8', 'cp1252', 'iso-8859-1', 'windows-1252']
        df = None
        for encoding in encodings:
            try:
                df = pd.read_csv(temp_file_path, encoding=encoding, nrows=10)
                break
            except Exception:
                continue
        if df is None:
            messages.error(request, 'Não foi possível ler o arquivo CSV. Tente salvar como CSV (separado por vírgula) e tente novamente.')
            return redirect('core:preview_import_clientes')
        colunas = list(df.columns)
        preview = df.head(5).values.tolist()
        
        # Lista de campos para mapeamento de clientes
        campos = ['nome', 'cpf_cnpj', 'rg', 'data_nascimento', 'telefone', 'email', 'endereco', 'cidade', 'estado', 'cep', 'tipo', 'observacoes']
        
        return render(request, 'core/preview_import_clientes.html', {
            'colunas': colunas,
            'preview': preview,
            'arquivo_nome': file.name,
            'campos': campos,
        })
    
    # Lista de campos para mapeamento (mesmo quando não há arquivo)
    campos = ['nome', 'cpf_cnpj', 'rg', 'data_nascimento', 'telefone', 'email', 'endereco', 'cidade', 'estado', 'cep', 'tipo', 'observacoes']
    return render(request, 'core/preview_import_clientes.html', {'campos': campos})

@csrf_exempt
@login_required
def preview_import_vendas(request):
    """Pré-visualização do CSV para importação de vendas"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Salvar o arquivo temporariamente
        import tempfile
        import os
        
        # Criar arquivo temporário
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_uploads')
        os.makedirs(temp_dir, exist_ok=True)
        
        temp_file_path = os.path.join(temp_dir, f"temp_vendas_{request.user.id}_{file.name}")
        with open(temp_file_path, 'wb+') as temp_file:
            for chunk in file.chunks():
                temp_file.write(chunk)
        
        # Armazenar o caminho do arquivo temporário na sessão
        request.session['temp_file_path_vendas'] = temp_file_path
        
        encodings = ['latin1', 'utf-8-sig', 'utf-8', 'cp1252', 'iso-8859-1', 'windows-1252']
        df = None
        for encoding in encodings:
            try:
                df = pd.read_csv(temp_file_path, encoding=encoding, nrows=10)
                break
            except Exception:
                continue
        if df is None:
            messages.error(request, 'Não foi possível ler o arquivo CSV. Tente salvar como CSV (separado por vírgula) e tente novamente.')
            return redirect('core:preview_import_vendas')
        colunas = list(df.columns)
        preview = df.head(5).values.tolist()
        
        # Lista de campos para mapeamento de vendas
        campos = ['moto_chassi', 'comprador_cpf', 'vendedor_username', 'loja_nome', 'origem', 'forma_pagamento', 'status', 'valor_venda', 'valor_entrada', 'comissao_vendedor', 'data_atendimento', 'data_venda', 'observacoes']
        
        return render(request, 'core/preview_import_vendas.html', {
            'colunas': colunas,
            'preview': preview,
            'arquivo_nome': file.name,
            'campos': campos,
        })
    
    # Lista de campos para mapeamento (mesmo quando não há arquivo)
    campos = ['moto_chassi', 'comprador_cpf', 'vendedor_username', 'loja_nome', 'origem', 'forma_pagamento', 'status', 'valor_venda', 'valor_entrada', 'comissao_vendedor', 'data_atendimento', 'data_venda', 'observacoes']
    return render(request, 'core/preview_import_vendas.html', {'campos': campos})

@login_required
def test_template(request):
    """View de teste para verificar se o template base está funcionando"""
    return render(request, 'core/test_template.html', {
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def usuario_menu_manage(request, usuario_id):
    """Gerencia os menus específicos de um usuário"""
    try:
        usuario = Usuario.objects.get(pk=usuario_id)
    except Usuario.DoesNotExist:
        messages.error(request, 'Usuário não encontrado.')
        return redirect('core:usuario_list')
    
    # Verificar permissões
    if not (request.user.is_superuser or 
            (hasattr(request.user, 'usuario_sistema') and 
             request.user.usuario_sistema.perfil.nome in ['admin', 'gerente', 'ti'])):
        return render(request, 'core/acesso_negado.html', {
            'mensagem': 'Você não tem permissão para acessar esta funcionalidade.'
        })
    
    if request.method == 'POST':
        # Processar o formulário
        modulos = [
            ('clientes', 'Clientes'),
            ('motocicletas', 'Motocicletas'),
            ('vendas', 'Vendas'),
            ('consignacoes', 'Consignações'),
            ('seguros', 'Seguros'),
            ('usuarios', 'Usuários'),
            ('lojas', 'Lojas'),
            ('controle_chaves', 'Controle de Chaves'),
            ('ocorrencias', 'Ocorrências'),
            ('seguradoras', 'Seguradoras'),
            ('bens', 'Bens'),
            ('cotacoes', 'Cotações'),
            ('financeiro', 'Financeiro'),
            ('pre_venda', 'Pré-Venda'),
        ]
        
        # Limpar configurações existentes do usuário
        usuario.menus_usuario.all().delete()
        
        # Criar novas configurações baseadas no formulário
        for cod, nome in modulos:
            if request.POST.get(f'{usuario.pk}_{cod}'):
                MenuUsuario.objects.create(
                    usuario=usuario,
                    modulo=cod,
                    ativo=True
                )
        
        messages.success(request, f'Menus do usuário {usuario.user.get_full_name()} atualizados com sucesso!')
        return redirect('core:usuario_list')
    
    # Preparar dados para o template
    modulos = [
        ('clientes', 'Clientes'),
        ('motocicletas', 'Motocicletas'),
        ('vendas', 'Vendas'),
        ('consignacoes', 'Consignações'),
        ('seguros', 'Seguros'),
        ('usuarios', 'Usuários'),
        ('lojas', 'Lojas'),
        ('controle_chaves', 'Controle de Chaves'),
        ('ocorrencias', 'Ocorrências'),
        ('seguradoras', 'Seguradoras'),
        ('bens', 'Bens'),
        ('cotacoes', 'Cotações'),
        ('financeiro', 'Financeiro'),
        ('pre_venda', 'Pré-Venda'),
    ]
    
    # Verificar quais módulos estão ativos para o usuário
    menus_usuario = {}
    for cod, nome in modulos:
        menus_usuario[cod] = usuario.modulo_ativo(cod)
    
    context = {
        'usuario': usuario,
        'modulos': modulos,
        'menus_usuario': menus_usuario,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/usuario_menu_manage.html', context)

@login_required
def import_motocicletas(request):
    """Processa a importação de motocicletas com base no mapeamento"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')

    if request.method == 'POST':
        # Verificar se existe arquivo temporário na sessão
        temp_file_path = request.session.get('temp_file_path')
        if not temp_file_path:
            messages.error(request, 'Arquivo não encontrado. Faça upload novamente.')
            return redirect('core:preview_import_motocicletas')
        
        # Verificar se o arquivo existe
        if not os.path.exists(temp_file_path):
            messages.error(request, 'Arquivo temporário não encontrado. Faça upload novamente.')
            # Limpar sessão
            if 'temp_file_path' in request.session:
                del request.session['temp_file_path']
            return redirect('core:preview_import_motocicletas')

        # Extrair mapeamento das colunas
        column_mapping = {}
        campos = ['marca', 'modelo', 'ano', 'ano_fabricacao', 'cor', 'placa', 'chassi', 'rodagem', 'valor_entrada', 'valor_atual', 'status', 'observacoes']
        
        for campo in campos:
            map_key = f'map_{campo}'
            if map_key in request.POST and request.POST[map_key]:
                # Converter para o formato esperado pelo importador (lista de possíveis nomes)
                column_mapping[campo] = [request.POST[map_key]]

        # Verificar campos obrigatórios - REMOVIDO PARA PERMITIR IMPORTAR MESMO COM DADOS DIVERGENTES
        # campos_obrigatorios = ['marca', 'modelo', 'chassi']
        # campos_faltando = [campo for campo in campos_obrigatorios if campo not in column_mapping]
        # 
        # if campos_faltando:
        #     messages.error(request, f'Campos obrigatórios não mapeados: {", ".join(campos_faltando)}')
        #     return redirect('core:preview_import_motocicletas')

        try:
            # Importar usando o importador
            from .importers import DataImporter
            importer = DataImporter()
            success = importer.import_motocicletas(temp_file_path, column_mapping)
            summary = importer.get_import_summary()
            
            # Limpar arquivo temporário
            try:
                os.remove(temp_file_path)
                del request.session['temp_file_path']
            except Exception as e:
                logger.warning(f"Erro ao limpar arquivo temporário: {str(e)}")

            if success:
                messages.success(request, f'Importação concluída! {summary["success_count"]} motocicletas importadas com sucesso.')
                if summary.get("skipped_count", 0) > 0:
                    messages.warning(request, f'{summary["skipped_count"]} motocicletas foram ignoradas (duplicadas ou com erros).')
                if summary["error_count"] > 0:
                    messages.warning(request, f'{summary["error_count"]} registros com erros foram ignorados.')
                
                # Armazenar relatório na sessão para exibição
                if summary.get("skipped_motos"):
                    request.session['import_report'] = {
                        'success_count': summary["success_count"],
                        'skipped_count': summary.get("skipped_count", 0),
                        'error_count': summary["error_count"],
                        'skipped_motos': summary["skipped_motos"][:50]  # Limitar a 50 para não sobrecarregar a sessão
                    }
            else:
                messages.error(request, f'Erro na importação: {summary.get("errors", ["Erro desconhecido"])[0] if summary.get("errors") else "Erro desconhecido"}')

        except Exception as e:
            logger.error(f"Erro durante a importação: {str(e)}")
            messages.error(request, f'Erro durante a importação: {str(e)}')
            # Limpar arquivo temporário em caso de erro
            try:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                if 'temp_file_path' in request.session:
                    del request.session['temp_file_path']
            except Exception as cleanup_error:
                logger.warning(f"Erro ao limpar arquivo temporário: {str(cleanup_error)}")

        return redirect('core:preview_import_motocicletas')

    return redirect('core:preview_import_motocicletas')

@login_required
def import_clientes_process(request):
    """Processa a importação de clientes com base no mapeamento"""
    if not request.user.is_superuser:
        messages.error(request, 'Apenas administradores podem importar dados.')
        return redirect('core:dashboard')

    if request.method == 'POST':
        # Verificar se existe arquivo temporário na sessão
        temp_file_path = request.session.get('temp_file_path_clientes')
        if not temp_file_path:
            messages.error(request, 'Arquivo não encontrado. Faça upload novamente.')
            return redirect('core:preview_import_clientes')
        
        # Verificar se o arquivo existe
        if not os.path.exists(temp_file_path):
            messages.error(request, 'Arquivo temporário não encontrado. Faça upload novamente.')
            # Limpar sessão
            if 'temp_file_path_clientes' in request.session:
                del request.session['temp_file_path_clientes']
            return redirect('core:preview_import_clientes')

        # Extrair mapeamento das colunas
        column_mapping = {}
        campos = ['nome', 'cpf_cnpj', 'rg', 'data_nascimento', 'telefone', 'email', 'endereco', 'cidade', 'estado', 'cep', 'tipo', 'observacoes']
        
        for campo in campos:
            map_key = f'map_{campo}'
            if map_key in request.POST and request.POST[map_key]:
                # Converter para o formato esperado pelo importador (lista de possíveis nomes)
                column_mapping[campo] = [request.POST[map_key]]

        try:
            # Importar usando o importador
            from .importers import DataImporter
            importer = DataImporter()
            success = importer.import_clientes(temp_file_path, column_mapping)
            summary = importer.get_import_summary()
            
            # Limpar arquivo temporário
            try:
                os.remove(temp_file_path)
                del request.session['temp_file_path_clientes']
            except Exception as e:
                logger.warning(f"Erro ao limpar arquivo temporário: {str(e)}")

            if success:
                messages.success(request, f'Importação concluída! {summary["success_count"]} clientes importados com sucesso.')
                if summary.get("skipped_count", 0) > 0:
                    messages.warning(request, f'{summary["skipped_count"]} clientes foram ignorados (duplicados ou com erros).')
                if summary["error_count"] > 0:
                    messages.warning(request, f'{summary["error_count"]} registros com erros foram ignorados.')
            else:
                messages.error(request, f'Erro na importação: {summary.get("errors", ["Erro desconhecido"])[0] if summary.get("errors") else "Erro desconhecido"}')

        except Exception as e:
            logger.error(f"Erro durante a importação: {str(e)}")
            messages.error(request, f'Erro durante a importação: {str(e)}')
            # Limpar arquivo temporário em caso de erro
            try:
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                if 'temp_file_path_clientes' in request.session:
                    del request.session['temp_file_path_clientes']
            except Exception as cleanup_error:
                logger.warning(f"Erro ao limpar arquivo temporário: {str(cleanup_error)}")

        return redirect('core:preview_import_clientes')

    return redirect('core:preview_import_clientes')

def try_read_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ['.xlsx', '.xls']:
        try:
            df = pd.read_excel(file_path)
            if df.shape[1] > 1:
                return df, 'excel', None
        except Exception as e:
            return None, 'excel', str(e)
    # Tenta CSV com diferentes separadores e encodings
    separators = [',', ';', '\t']
    encodings = ['utf-8', 'latin1']
    for sep in separators:
        for enc in encodings:
            try:
                df = pd.read_csv(file_path, sep=sep, encoding=enc)
                if df.shape[1] > 1:
                    return df, sep, enc
            except Exception as e:
                continue
    return None, None, None

@login_required
def cliente_reactivate(request, pk):
    """Reativa um cliente inativo"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_cliente')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para reativar clientes.'})
    
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.ativo = True
        cliente.save()
        messages.success(request, f'Cliente {cliente.nome} reativado com sucesso!')
        return redirect('core:cliente_detail', pk=cliente.pk)
    
    return render(request, 'core/cliente_reactivate.html', {
        'cliente': cliente,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def motocicleta_remove_proprietario(request, pk):
    """Remove o relacionamento entre motocicleta e proprietário"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_motocicleta')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para editar motocicletas.'})
    
    motocicleta = get_object_or_404(Motocicleta, pk=pk)
    
    if request.method == 'POST':
        # Verificar se há vendas registradas
        if motocicleta.vendas.exists():
            messages.error(request, 'Não é possível remover o proprietário de uma motocicleta que possui vendas registradas.')
            return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        
        # Verificar se tem consignação ativa
        try:
            if motocicleta.consignacao and motocicleta.consignacao.status in ['disponivel', 'em_negociacao']:
                messages.error(request, 'Não é possível remover o proprietário de uma motocicleta que possui consignação ativa.')
                return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        except Exception:
            pass
        
        # Verificar se tem seguro ativo
        try:
            if motocicleta.seguro and motocicleta.seguro.status == 'ativo':
                messages.error(request, 'Não é possível remover o proprietário de uma motocicleta que possui seguro ativo.')
                return redirect('core:motocicleta_detail', pk=motocicleta.pk)
        except Exception:
            pass
        
        # Registrar no histórico se houver proprietário atual
        if motocicleta.proprietario:
            from .models import HistoricoProprietario
            HistoricoProprietario.objects.create(
                moto=motocicleta,
                proprietario=motocicleta.proprietario,
                data_inicio=motocicleta.data_entrada,
                data_fim=timezone.now().date(),
                motivo='remocao',
                valor_transacao=motocicleta.valor_atual
            )
        
        # Remover o proprietário
        proprietario_anterior = motocicleta.proprietario
        motocicleta.proprietario = None
        motocicleta.save()
        
        messages.success(request, f'Proprietário removido com sucesso da motocicleta {motocicleta}.')
        if proprietario_anterior:
            messages.info(request, f'Histórico de propriedade registrado para {proprietario_anterior.nome}.')
        
        return redirect('core:motocicleta_detail', pk=motocicleta.pk)
    
    return render(request, 'core/motocicleta_remove_proprietario.html', {
        'motocicleta': motocicleta,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def motocicleta_transferir_propriedade(request, pk):
    """Transfere a propriedade da motocicleta para outro cliente"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_motocicleta')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para transferir propriedade de motocicletas.'})
    
    motocicleta = get_object_or_404(Motocicleta, pk=pk)
    from .models import Cliente, HistoricoProprietario
    if request.method == 'POST':
        novo_proprietario_id = request.POST.get('novo_proprietario')
        if not novo_proprietario_id:
            messages.error(request, 'Selecione o novo proprietário.')
            return redirect('core:motocicleta_transferir_propriedade', pk=pk)
        novo_proprietario = get_object_or_404(Cliente, pk=novo_proprietario_id)
        # Registrar histórico do proprietário anterior
        if motocicleta.proprietario:
            HistoricoProprietario.objects.create(
                moto=motocicleta,
                proprietario=motocicleta.proprietario,
                data_inicio=motocicleta.data_entrada,
                data_fim=timezone.now().date(),
                motivo='transferencia',
                valor_transacao=motocicleta.valor_atual
            )
        # Transferir propriedade
        motocicleta.proprietario = novo_proprietario
        motocicleta.save()
        messages.success(request, f'Propriedade transferida para {novo_proprietario.nome} com sucesso!')
        return redirect('core:motocicleta_detail', pk=pk)
    # Listar clientes ativos para seleção
    clientes = Cliente.objects.filter(ativo=True).order_by('nome')
    return render(request, 'core/motocicleta_transferir_propriedade.html', {
        'motocicleta': motocicleta,
        'clientes': clientes,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

# ============================================================================
# VIEWS DO MÓDULO FINANCEIRO
# ============================================================================

@login_required
def dashboard_financeiro(request):
    """Dashboard financeiro com KPIs, gráficos e análises"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_vendafinanceira')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para acessar o dashboard financeiro.'})
    
    # Obter parâmetros de filtro
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    loja_id = request.GET.get('loja')
    
    # Converter datas
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    else:
        data_inicio = timezone.now().date() - timedelta(days=30)
    
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    else:
        data_fim = timezone.now().date()
    
    # Obter loja se especificada
    loja = None
    if loja_id:
        try:
            loja = Loja.objects.get(id=loja_id)
        except Loja.DoesNotExist:
            pass
    
    # Calcular dados financeiros
    from .financeiro import CalculadoraFinanceira
    
    calc = CalculadoraFinanceira(loja=loja, data_inicio=data_inicio, data_fim=data_fim)
    resumo = calc.get_resumo_financeiro()
    
    # Obter dados para tabelas
    from .models import Pagamento, VendaFinanceira
    
    # Pagamentos pendentes
    pagamentos_pendentes = Pagamento.objects.filter(
        pago=False,
        loja=loja if loja else Q(),
        vencimento__lte=data_fim + timedelta(days=30)  # Próximos 30 dias
    ).order_by('vencimento')[:10]
    
    # Vendas com maior lucro (ordenadas por data, depois calculamos o lucro em Python)
    if loja:
        vendas_maior_lucro = Venda.objects.filter(
            data_venda__gte=data_inicio,
            data_venda__lte=data_fim,
            loja=loja
        ).order_by('-data_venda')[:50]  # Pegamos mais registros para depois filtrar os top 10
    else:
        vendas_maior_lucro = Venda.objects.filter(
            data_venda__gte=data_inicio,
            data_venda__lte=data_fim
        ).order_by('-data_venda')[:50]  # Pegamos mais registros para depois filtrar os top 10
    
    # Calcular lucro para cada venda e ordenar
    vendas_com_lucro = []
    for venda in vendas_maior_lucro:
        lucro = venda.valor_venda - (venda.valor_entrada or 0)
        vendas_com_lucro.append((venda, lucro))
    
    # Ordenar por lucro e pegar top 10
    vendas_com_lucro.sort(key=lambda x: x[1], reverse=True)
    vendas_maior_lucro = [venda for venda, lucro in vendas_com_lucro[:10]]
    
    # Listar todas as lojas para o filtro
    lojas = Loja.objects.filter(ativo=True).order_by('nome')
    
    context = {
        'resumo': resumo,
        'alertas': resumo['alertas'],
        'pagamentos_pendentes': pagamentos_pendentes,
        'vendas_maior_lucro': vendas_maior_lucro,
        'lojas': lojas,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'loja_selecionada': loja,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/dashboard_financeiro.html', context)

@login_required
def venda_financeira_list(request):
    """Lista de vendas financeiras - Redireciona para lista de vendas normais"""
    messages.info(request, 'Visualizando lista de vendas de motocicletas.')
    return redirect('core:venda_list')

@login_required
def venda_financeira_create(request):
    """Criar nova venda de moto - Redireciona para formulário de vendas"""
    messages.info(request, 'Use o formulário de vendas de motocicletas para registrar vendas. O módulo financeiro é para análise e relatórios.')
    return redirect('core:venda_create')

@login_required
def venda_financeira_detail(request, pk):
    """Detalhes da venda financeira - Redireciona para detalhes da venda normal"""
    messages.info(request, 'Visualizando detalhes da venda de motocicleta.')
    return redirect('core:venda_detail', pk=pk)

@login_required
def despesa_list(request):
    """Lista de despesas"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_despesa')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar despesas.'})
    
    despesas = Despesa.objects.all().order_by('-data')
    
    # Filtros
    categoria = request.GET.get('categoria')
    if categoria:
        despesas = despesas.filter(categoria=categoria)
    
    data_inicio = request.GET.get('data_inicio')
    if data_inicio:
        despesas = despesas.filter(data__gte=data_inicio)
    
    data_fim = request.GET.get('data_fim')
    if data_fim:
        despesas = despesas.filter(data__lte=data_fim)
    
    # Cálculos
    total_valor = despesas.aggregate(total=Sum('valor'))['total'] or 0
    media_valor = total_valor / despesas.count() if despesas.count() > 0 else 0
    
    context = {
        'despesas': despesas,
        'total_valor': total_valor,
        'media_valor': media_valor,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/despesa_list.html', context)

@login_required
def despesa_create(request):
    """Criar nova despesa"""
    if not (request.user.is_superuser or request.user.has_perm('core.add_despesa')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para criar despesas.'})
    
    from .forms import DespesaForm
    
    if request.method == 'POST':
        form = DespesaForm(request.POST)
        if form.is_valid():
            despesa = form.save(commit=False)
            despesa.responsavel = request.user.usuario_sistema
            despesa.save()
            messages.success(request, 'Despesa criada com sucesso!')
            return redirect('core:despesa_detail', pk=despesa.pk)
    else:
        form = DespesaForm()
    
    context = {
        'form': form,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/despesa_form.html', context)

@login_required
def despesa_detail(request, pk):
    """Detalhes da despesa"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_despesa')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar despesas.'})
    
    despesa = get_object_or_404(Despesa, pk=pk)
    
    context = {
        'despesa': despesa,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/despesa_detail.html', context)

@login_required
def despesa_update(request, pk):
    """Editar despesa"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_despesa')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para editar despesas.'})
    
    despesa = get_object_or_404(Despesa, pk=pk)
    from .forms import DespesaForm
    
    if request.method == 'POST':
        form = DespesaForm(request.POST, instance=despesa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Despesa atualizada com sucesso!')
            return redirect('core:despesa_detail', pk=despesa.pk)
    else:
        form = DespesaForm(instance=despesa)
    
    context = {
        'form': form,
        'despesa': despesa,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/despesa_form.html', context)

@login_required
def despesa_delete(request, pk):
    """Excluir despesa"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_despesa')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir despesas.'})
    
    despesa = get_object_or_404(Despesa, pk=pk)
    
    if request.method == 'POST':
        despesa.delete()
        messages.success(request, 'Despesa excluída com sucesso!')
        return redirect('core:despesa_list')
    
    context = {
        'despesa': despesa,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/despesa_confirm_delete.html', context)

@login_required
def receita_extra_list(request):
    """Lista de receitas extras"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_receitaextra')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar receitas extras.'})
    
    receitas = ReceitaExtra.objects.all().order_by('-data')
    
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    if data_inicio:
        receitas = receitas.filter(data__gte=data_inicio)
    
    data_fim = request.GET.get('data_fim')
    if data_fim:
        receitas = receitas.filter(data__lte=data_fim)
    
    # Cálculos
    total_receitas = receitas.aggregate(total=Sum('valor'))['total'] or 0
    media_receitas = total_receitas / receitas.count() if receitas.count() > 0 else 0
    
    context = {
        'receitas': receitas,
        'total_receitas': total_receitas,
        'media_receitas': media_receitas,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/receita_extra_list.html', context)

@login_required
def receita_extra_create(request):
    """Criar nova receita extra"""
    if not (request.user.is_superuser or request.user.has_perm('core.add_receitaextra')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para criar receitas extras.'})
    
    from .forms import ReceitaExtraForm
    
    if request.method == 'POST':
        form = ReceitaExtraForm(request.POST)
        if form.is_valid():
            receita = form.save(commit=False)
            receita.responsavel = request.user.usuario_sistema
            receita.save()
            messages.success(request, 'Receita extra criada com sucesso!')
            return redirect('core:receita_extra_detail', pk=receita.pk)
    else:
        form = ReceitaExtraForm()
    
    context = {
        'form': form,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/receita_extra_form.html', context)

@login_required
def receita_extra_detail(request, pk):
    """Detalhes da receita extra"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_receitaextra')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar receitas extras.'})
    
    receita = get_object_or_404(ReceitaExtra, pk=pk)
    
    context = {
        'receita': receita,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/receita_extra_detail.html', context)

@login_required
def receita_extra_update(request, pk):
    """Editar receita extra"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_receitaextra')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para editar receitas extras.'})
    
    receita = get_object_or_404(ReceitaExtra, pk=pk)
    from .forms import ReceitaExtraForm
    
    if request.method == 'POST':
        form = ReceitaExtraForm(request.POST, instance=receita)
        if form.is_valid():
            form.save()
            messages.success(request, 'Receita extra atualizada com sucesso!')
            return redirect('core:receita_extra_detail', pk=receita.pk)
    else:
        form = ReceitaExtraForm(instance=receita)
    
    context = {
        'form': form,
        'receita': receita,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/receita_extra_form.html', context)

@login_required
def receita_extra_delete(request, pk):
    """Excluir receita extra"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_receitaextra')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir receitas extras.'})
    
    receita = get_object_or_404(ReceitaExtra, pk=pk)
    
    if request.method == 'POST':
        receita.delete()
        messages.success(request, 'Receita extra excluída com sucesso!')
        return redirect('core:receita_extra_list')
    
    context = {
        'receita': receita,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/receita_extra_confirm_delete.html', context)

@login_required
def pagamento_list(request):
    """Lista de pagamentos"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_pagamento')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar pagamentos.'})
    
    from django.db.models import Sum
    from django.utils import timezone
    
    pagamentos = Pagamento.objects.all().order_by('vencimento')
    
    # Filtros
    tipo = request.GET.get('tipo')
    if tipo:
        pagamentos = pagamentos.filter(tipo=tipo)
    
    status = request.GET.get('status')
    if status == 'pago':
        pagamentos = pagamentos.filter(pago=True)
    elif status == 'pendente':
        pagamentos = pagamentos.filter(pago=False)
    elif status == 'atrasado':
        pagamentos = pagamentos.filter(pago=False, vencimento__lt=timezone.now().date())
    
    # Cálculos
    total_pago = pagamentos.filter(pago=True).aggregate(total=Sum('valor'))['total'] or 0
    total_pendente = pagamentos.filter(pago=False).aggregate(total=Sum('valor'))['total'] or 0
    vencidos_count = pagamentos.filter(pago=False, vencimento__lt=timezone.now().date()).count()
    
    # Lojas para filtro
    from .models import Loja
    lojas = Loja.objects.all()
    
    context = {
        'pagamentos': pagamentos,
        'total_pago': total_pago,
        'total_pendente': total_pendente,
        'vencidos_count': vencidos_count,
        'lojas': lojas,
        'hoje': timezone.now().date(),
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/pagamento_list.html', context)

@login_required
def pagamento_create(request):
    """Criar novo pagamento"""
    if not (request.user.is_superuser or request.user.has_perm('core.add_pagamento')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para criar pagamentos.'})
    
    from .forms import PagamentoForm
    
    if request.method == 'POST':
        form = PagamentoForm(request.POST)
        if form.is_valid():
            pagamento = form.save(commit=False)
            pagamento.responsavel = request.user.usuario_sistema
            pagamento.save()
            messages.success(request, 'Pagamento criado com sucesso!')
            return redirect('core:pagamento_detail', pk=pagamento.pk)
    else:
        form = PagamentoForm()
    
    context = {
        'form': form,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/pagamento_form.html', context)

@login_required
def pagamento_detail(request, pk):
    """Detalhes do pagamento"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_pagamento')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar pagamentos.'})
    
    pagamento = get_object_or_404(Pagamento, pk=pk)
    
    context = {
        'pagamento': pagamento,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/pagamento_detail.html', context)

@login_required
def pagamento_update(request, pk):
    """Editar pagamento"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_pagamento')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para editar pagamentos.'})
    
    pagamento = get_object_or_404(Pagamento, pk=pk)
    from .forms import PagamentoForm
    
    if request.method == 'POST':
        form = PagamentoForm(request.POST, instance=pagamento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pagamento atualizado com sucesso!')
            return redirect('core:pagamento_detail', pk=pagamento.pk)
    else:
        form = PagamentoForm(instance=pagamento)
    
    context = {
        'form': form,
        'pagamento': pagamento,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/pagamento_form.html', context)

@login_required
def pagamento_delete(request, pk):
    """Excluir pagamento"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_pagamento')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir pagamentos.'})
    
    pagamento = get_object_or_404(Pagamento, pk=pk)
    
    if request.method == 'POST':
        pagamento.delete()
        messages.success(request, 'Pagamento excluído com sucesso!')
        return redirect('core:pagamento_list')
    
    context = {
        'pagamento': pagamento,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    
    return render(request, 'core/pagamento_confirm_delete.html', context)

@login_required
def exportar_dashboard_financeiro_xlsx(request):
    """Exporta os dados do dashboard financeiro em formato XLSX"""
    from .financeiro import CalculadoraFinanceira
    from .models import Motocicleta, Venda, Despesa, Pagamento, ReceitaExtra
    from datetime import datetime, timedelta
    from django.utils import timezone

    def parse_data(data_str, default):
        if not data_str:
            return default
        try:
            # Tenta formato ISO
            return datetime.strptime(data_str, '%Y-%m-%d').date()
        except Exception:
            pass
        try:
            # Tenta formato por extenso em português
            meses = {'janeiro':1, 'fevereiro':2, 'março':3, 'marco':3, 'abril':4, 'maio':5, 'junho':6, 'julho':7, 'agosto':8, 'setembro':9, 'outubro':10, 'novembro':11, 'dezembro':12}
            partes = data_str.lower().replace('de ', '').split()
            if len(partes) == 3:
                dia = int(partes[0])
                mes = meses.get(partes[1], 1)
                ano = int(partes[2])
                return datetime(ano, mes, dia).date()
        except Exception:
            pass
        return default

    data_inicio = parse_data(request.GET.get('data_inicio'), timezone.now().date() - timedelta(days=30))
    data_fim = parse_data(request.GET.get('data_fim'), timezone.now().date())
    loja_id = request.GET.get('loja')

    loja = None
    if loja_id:
        try:
            from .models import Loja
            loja = Loja.objects.get(id=loja_id)
        except Loja.DoesNotExist:
            pass

    calc = CalculadoraFinanceira(loja=loja, data_inicio=data_inicio, data_fim=data_fim)
    resumo = calc.get_resumo_financeiro()

    # Workbook
    wb = openpyxl.Workbook()
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'

    # Resumo dos cards
    ws_resumo.append(['Indicador', 'Valor'])
    ws_resumo.append(['Receita Bruta', float(resumo['receita_bruta'])])
    ws_resumo.append(['Lucro Líquido', float(resumo['lucro_total'])])
    ws_resumo.append(['Margem de Lucro (%)', float(resumo['margem_lucro'])])
    ws_resumo.append(['Projeção de Caixa', float(resumo['projecao_caixa'])])
    ws_resumo.append(['Motos em Estoque', resumo['total_motos_estoque']])
    ws_resumo.append(['0KM', resumo['total_0km']])
    ws_resumo.append(['Usadas', resumo['total_usadas']])
    ws_resumo.append(['Consignação', resumo['total_consignacao']])
    ws_resumo.append(['Valor em Estoque', float(resumo['valor_estoque'])])
    ws_resumo.append(['Total Gasto em Motos', float(resumo['valor_gasto_motos'])])

    for cell in ws_resumo[1]:
        cell.font = Font(bold=True)

    # Motos em estoque
    ws_motos = wb.create_sheet('Motos em Estoque')
    ws_motos.append(['Marca', 'Modelo', 'Ano', 'Tipo', 'Status', 'Valor Atual', 'Valor Entrada', 'Data Entrada'])
    for cell in ws_motos[1]:
        cell.font = Font(bold=True)
    motos_estoque = Motocicleta.objects.filter(status='estoque', ativo=True, **(calc.get_filtros_loja()))
    for m in motos_estoque:
        ws_motos.append([
            m.marca, m.modelo, m.ano, m.get_tipo_entrada_display(), m.get_status_display(),
            float(m.valor_atual or 0), float(m.valor_entrada or 0), m.data_entrada.strftime('%d/%m/%Y') if m.data_entrada else ''
        ])

    # Vendas do período
    ws_vendas = wb.create_sheet('Vendas')
    ws_vendas.append(['ID', 'Data', 'Cliente', 'Motocicleta', 'Valor Venda', 'Valor Entrada', 'Lucro'])
    for cell in ws_vendas[1]:
        cell.font = Font(bold=True)
    vendas = Venda.objects.filter(data_venda__gte=data_inicio, data_venda__lte=data_fim)
    if loja:
        vendas = vendas.filter(loja=loja)
    for v in vendas:
        lucro = (v.valor_venda or 0) - (v.valor_entrada or 0)
        ws_vendas.append([
            v.id, v.data_venda.strftime('%d/%m/%Y') if v.data_venda else '',
            v.comprador.nome if v.comprador else '',
            f'{v.moto.marca} {v.moto.modelo}' if v.moto else '',
            float(v.valor_venda or 0), float(v.valor_entrada or 0), float(lucro)
        ])

    # Despesas do período
    ws_despesas = wb.create_sheet('Despesas')
    ws_despesas.append(['ID', 'Data', 'Descrição', 'Categoria', 'Valor', 'Loja', 'Responsável'])
    for cell in ws_despesas[1]:
        cell.font = Font(bold=True)
    despesas = Despesa.objects.filter(data__gte=data_inicio, data__lte=data_fim)
    if loja:
        despesas = despesas.filter(loja=loja)
    for d in despesas:
        ws_despesas.append([
            d.id, d.data.strftime('%d/%m/%Y') if d.data else '', d.descricao, d.get_categoria_display(),
            float(d.valor or 0), d.loja.nome if d.loja else '', d.responsavel.user.get_full_name() if d.responsavel else ''
        ])

    # Pagamentos pendentes
    ws_pagamentos = wb.create_sheet('Pagamentos Pendentes')
    ws_pagamentos.append(['ID', 'Tipo', 'Referente', 'Valor', 'Vencimento', 'Loja', 'Responsável'])
    for cell in ws_pagamentos[1]:
        cell.font = Font(bold=True)
    pagamentos = Pagamento.objects.filter(pago=False, loja=loja if loja else None)
    for p in pagamentos:
        ws_pagamentos.append([
            p.id, p.get_tipo_display(), p.get_referente_a_display(), float(p.valor or 0),
            p.vencimento.strftime('%d/%m/%Y') if p.vencimento else '',
            p.loja.nome if p.loja else '', p.responsavel.user.get_full_name() if p.responsavel else ''
        ])

    # Receitas extras
    ws_receitas = wb.create_sheet('Receitas Extras')
    ws_receitas.append(['ID', 'Data', 'Descrição', 'Valor', 'Loja', 'Responsável'])
    for cell in ws_receitas[1]:
        cell.font = Font(bold=True)
    receitas = ReceitaExtra.objects.filter(data__gte=data_inicio, data__lte=data_fim)
    if loja:
        receitas = receitas.filter(loja=loja)
    for r in receitas:
        ws_receitas.append([
            r.id, r.data.strftime('%d/%m/%Y') if r.data else '', r.descricao, float(r.valor or 0),
            r.loja.nome if r.loja else '', r.responsavel.user.get_full_name() if r.responsavel else ''
        ])

    # Ajustar largura das colunas
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # Resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dashboard_financeiro.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_vendas_xlsx(request):
    """Exporta lista de vendas em formato XLSX"""
    from .models import Venda
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Filtros
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    loja_id = request.GET.get('loja', '')
    
    vendas = Venda.objects.all()
    
    if status:
        vendas = vendas.filter(status=status)
    if data_inicio:
        vendas = vendas.filter(data_venda__gte=data_inicio)
    if data_fim:
        vendas = vendas.filter(data_venda__lte=data_fim)
    if loja_id:
        vendas = vendas.filter(loja_id=loja_id)
    
    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendas'
    
    # Cabeçalhos
    headers = ['ID', 'Data Venda', 'Comprador', 'Motocicleta', 'Vendedor', 'Loja', 'Valor Venda', 'Valor Entrada', 'Lucro', 'Status', 'Forma Pagamento']
    ws.append(headers)
    
    # Estilo do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Dados
    for venda in vendas:
        lucro = (venda.valor_venda or 0) - (venda.valor_entrada or 0)
        ws.append([
            venda.id,
            venda.data_venda.strftime('%d/%m/%Y') if venda.data_venda else '',
            venda.comprador.nome if venda.comprador else '',
            f'{venda.moto.marca} {venda.moto.modelo}' if venda.moto else '',
            venda.vendedor.user.get_full_name() if venda.vendedor else '',
            venda.loja.nome if venda.loja else '',
            float(venda.valor_venda or 0),
            float(venda.valor_entrada or 0),
            float(lucro),
            venda.get_status_display(),
            venda.get_forma_pagamento_display()
        ])
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vendas.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_despesas_xlsx(request):
    """Exporta lista de despesas em formato XLSX"""
    from .models import Despesa
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Filtros
    categoria = request.GET.get('categoria', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    loja_id = request.GET.get('loja', '')
    
    despesas = Despesa.objects.all()
    
    if categoria:
        despesas = despesas.filter(categoria=categoria)
    if data_inicio:
        despesas = despesas.filter(data__gte=data_inicio)
    if data_fim:
        despesas = despesas.filter(data__lte=data_fim)
    if loja_id:
        despesas = despesas.filter(loja_id=loja_id)
    
    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Despesas'
    
    # Cabeçalhos
    headers = ['ID', 'Data', 'Descrição', 'Categoria', 'Valor', 'Tipo', 'Centro Custo', 'Loja', 'Responsável']
    ws.append(headers)
    
    # Estilo do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Dados
    for despesa in despesas:
        ws.append([
            despesa.id,
            despesa.data.strftime('%d/%m/%Y') if despesa.data else '',
            despesa.descricao,
            despesa.get_categoria_display(),
            float(despesa.valor or 0),
            despesa.get_fixa_variavel_display(),
            despesa.centro_custo or '',
            despesa.loja.nome if despesa.loja else '',
            despesa.responsavel.user.get_full_name() if despesa.responsavel else ''
        ])
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=despesas.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_receitas_extras_xlsx(request):
    """Exporta lista de receitas extras em formato XLSX"""
    from .models import ReceitaExtra
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Filtros
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    loja_id = request.GET.get('loja', '')
    
    receitas = ReceitaExtra.objects.all()
    
    if data_inicio:
        receitas = receitas.filter(data__gte=data_inicio)
    if data_fim:
        receitas = receitas.filter(data__lte=data_fim)
    if loja_id:
        receitas = receitas.filter(loja_id=loja_id)
    
    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Receitas Extras'
    
    # Cabeçalhos
    headers = ['ID', 'Data', 'Descrição', 'Valor', 'Loja', 'Responsável', 'Observações']
    ws.append(headers)
    
    # Estilo do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Dados
    for receita in receitas:
        ws.append([
            receita.id,
            receita.data.strftime('%d/%m/%Y') if receita.data else '',
            receita.descricao,
            float(receita.valor or 0),
            receita.loja.nome if receita.loja else '',
            receita.responsavel.user.get_full_name() if receita.responsavel else '',
            receita.observacoes or ''
        ])
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=receitas_extras.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_pagamentos_xlsx(request):
    """Exporta lista de pagamentos em formato XLSX"""
    from .models import Pagamento
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Filtros
    tipo = request.GET.get('tipo', '')
    pago = request.GET.get('pago', '')
    loja_id = request.GET.get('loja', '')
    
    pagamentos = Pagamento.objects.all()
    
    if tipo:
        pagamentos = pagamentos.filter(tipo=tipo)
    if pago:
        pagamentos = pagamentos.filter(pago=pago == 'true')
    if loja_id:
        pagamentos = pagamentos.filter(loja_id=loja_id)
    
    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pagamentos'
    
    # Cabeçalhos
    headers = ['ID', 'Tipo', 'Referente', 'Valor', 'Vencimento', 'Pago', 'Data Pagamento', 'Loja', 'Responsável']
    ws.append(headers)
    
    # Estilo do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Dados
    for pagamento in pagamentos:
        ws.append([
            pagamento.id,
            pagamento.get_tipo_display(),
            pagamento.get_referente_a_display(),
            float(pagamento.valor or 0),
            pagamento.vencimento.strftime('%d/%m/%Y') if pagamento.vencimento else '',
            'Sim' if pagamento.pago else 'Não',
            pagamento.data_pagamento.strftime('%d/%m/%Y') if pagamento.data_pagamento else '',
            pagamento.loja.nome if pagamento.loja else '',
            pagamento.responsavel.user.get_full_name() if pagamento.responsavel else ''
        ])
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pagamentos.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_motocicletas_xlsx(request):
    """Exporta lista de motocicletas em formato XLSX"""
    from .models import Motocicleta
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Filtros
    status = request.GET.get('status', '')
    tipo_entrada = request.GET.get('tipo_entrada', '')
    marca = request.GET.get('marca', '')
    
    motocicletas = Motocicleta.objects.filter(ativo=True)
    
    if status:
        motocicletas = motocicletas.filter(status=status)
    if tipo_entrada:
        motocicletas = motocicletas.filter(tipo_entrada=tipo_entrada)
    if marca:
        motocicletas = motocicletas.filter(marca__icontains=marca)
    
    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Motocicletas'
    
    # Cabeçalhos
    headers = ['ID', 'Chassi', 'Placa', 'Marca', 'Modelo', 'Ano', 'Cor', 'Tipo Entrada', 'Status', 'Valor Atual', 'Valor Entrada', 'Data Entrada', 'Proprietário']
    ws.append(headers)
    
    # Estilo do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Dados
    for moto in motocicletas:
        ws.append([
            moto.id,
            moto.chassi,
            moto.placa or '',
            moto.marca,
            moto.modelo,
            moto.ano,
            moto.cor,
            moto.get_tipo_entrada_display(),
            moto.get_status_display(),
            float(moto.valor_atual or 0),
            float(moto.valor_entrada or 0),
            moto.data_entrada.strftime('%d/%m/%Y') if moto.data_entrada else '',
            moto.proprietario.nome if moto.proprietario else ''
        ])
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=motocicletas.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_clientes_xlsx(request):
    """Exporta lista de clientes em formato XLSX"""
    from .models import Cliente
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Filtros
    tipo = request.GET.get('tipo', '')
    ativo = request.GET.get('ativo', '')
    
    clientes = Cliente.objects.all()
    
    if tipo:
        clientes = clientes.filter(tipo=tipo)
    if ativo:
        clientes = clientes.filter(ativo=ativo == 'true')
    
    # Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    
    # Cabeçalhos
    headers = ['ID', 'Nome', 'CPF/CNPJ', 'RG', 'Data Nascimento', 'Telefone', 'Email', 'Tipo', 'Cidade', 'Estado', 'Ativo', 'Data Cadastro']
    ws.append(headers)
    
    # Estilo do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Dados
    for cliente in clientes:
        ws.append([
            cliente.id,
            cliente.nome,
            cliente.cpf_cnpj,
            cliente.rg or '',
            cliente.data_nascimento.strftime('%d/%m/%Y') if cliente.data_nascimento else '',
            cliente.telefone,
            cliente.email or '',
            cliente.get_tipo_display(),
            cliente.cidade or '',
            cliente.estado or '',
            'Sim' if cliente.ativo else 'Não',
            cliente.data_cadastro.strftime('%d/%m/%Y')
        ])
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'
    wb.save(response)
    return response

# ============================================================================
# VIEWS DE COMUNICAÇÃO DE VENDA
# ============================================================================

@login_required
def comunicacao_venda_list(request):
    """Lista de comunicações de venda com filtros"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar comunicações de venda.'})
    
    # Obter parâmetros de filtro
    status = request.GET.get('status', '')
    tipo = request.GET.get('tipo', '')
    venda_id = request.GET.get('venda', '')
    atrasadas = request.GET.get('atrasadas', '')
    
    # Query base
    comunicacoes = ComunicacaoVenda.objects.all().order_by('-data_criacao')
    
    # Aplicar filtros
    if status:
        comunicacoes = comunicacoes.filter(status=status)
    
    if tipo:
        comunicacoes = comunicacoes.filter(tipo=tipo)
    
    if venda_id:
        comunicacoes = comunicacoes.filter(venda_id=venda_id)
    
    if atrasadas == 'true':
        comunicacoes = [com for com in comunicacoes if com.atrasada]
    
    context = {
        'comunicacoes': comunicacoes,
        'status': status,
        'tipo': tipo,
        'venda_id': venda_id,
        'atrasadas': atrasadas,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    }
    return render(request, 'core/comunicacao_venda_list.html', context)

@login_required
def comunicacao_venda_create(request, venda_id=None):
    """Criação de nova comunicação de venda"""
    if not (request.user.is_superuser or request.user.has_perm('core.add_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para criar comunicações de venda.'})
    
    venda = None
    if venda_id:
        venda = get_object_or_404(Venda, pk=venda_id)
    
    if request.method == 'POST':
        form = ComunicacaoVendaForm(request.POST, venda=venda)
        if form.is_valid():
            comunicacao = form.save(commit=False)
            comunicacao.venda = venda
            comunicacao.responsavel = request.user.usuario_sistema
            comunicacao.save()
            messages.success(request, 'Comunicação de venda registrada com sucesso!')
            return redirect('core:comunicacao_venda_list')
    else:
        form = ComunicacaoVendaForm(venda=venda)
    
    return render(request, 'core/comunicacao_venda_form.html', {
        'form': form,
        'venda': venda,
        'usuario_sistema': request.user
    })

@login_required
def comunicacao_venda_update(request, pk):
    """Atualização de comunicação de venda"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para editar comunicações de venda.'})
    
    comunicacao = get_object_or_404(ComunicacaoVenda, pk=pk)
    if request.method == 'POST':
        form = ComunicacaoVendaForm(request.POST, instance=comunicacao, venda=comunicacao.venda)
        if form.is_valid():
            form.save()
            messages.success(request, 'Comunicação de venda atualizada com sucesso!')
            return redirect('core:comunicacao_venda_list')
    else:
        form = ComunicacaoVendaForm(instance=comunicacao, venda=comunicacao.venda)
    
    return render(request, 'core/comunicacao_venda_form.html', {
        'form': form,
        'comunicacao': comunicacao,
        'venda': comunicacao.venda,
        'usuario_sistema': request.user
    })

@login_required
def comunicacao_venda_detail(request, pk):
    """Detalhes da comunicação de venda"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar comunicações de venda.'})
    
    comunicacao = get_object_or_404(ComunicacaoVenda, pk=pk)
    return render(request, 'core/comunicacao_venda_detail.html', {
        'comunicacao': comunicacao,
        'usuario_sistema': request.user
    })

@login_required
def comunicacao_venda_delete(request, pk):
    """Exclusão de comunicação de venda"""
    if not (request.user.is_superuser or request.user.has_perm('core.delete_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para excluir comunicações de venda.'})
    
    comunicacao = get_object_or_404(ComunicacaoVenda, pk=pk)
    if request.method == 'POST':
        comunicacao.delete()
        messages.success(request, 'Comunicação de venda excluída com sucesso!')
        return redirect('core:comunicacao_venda_list')
    
    return render(request, 'core/comunicacao_venda_confirm_delete.html', {
        'comunicacao': comunicacao,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def comunicacao_venda_marcar_enviada(request, pk):
    """Marca uma comunicação como enviada"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para marcar comunicações como enviadas.'})
    
    comunicacao = get_object_or_404(ComunicacaoVenda, pk=pk)
    if request.method == 'POST':
        comunicacao.marcar_como_enviada()
        messages.success(request, 'Comunicação marcada como enviada com sucesso!')
        return redirect('core:comunicacao_venda_detail', pk=pk)
    
    return render(request, 'core/comunicacao_venda_confirm_enviada.html', {
        'comunicacao': comunicacao,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def comunicacao_venda_marcar_confirmada(request, pk):
    """Marca uma comunicação como confirmada"""
    if not (request.user.is_superuser or request.user.has_perm('core.change_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para marcar comunicações como confirmadas.'})
    
    comunicacao = get_object_or_404(ComunicacaoVenda, pk=pk)
    if request.method == 'POST':
        comunicacao.marcar_como_confirmada()
        messages.success(request, 'Comunicação marcada como confirmada com sucesso!')
        return redirect('core:comunicacao_venda_detail', pk=pk)
    
    return render(request, 'core/comunicacao_venda_confirm_confirmada.html', {
        'comunicacao': comunicacao,
        'usuario_sistema': getattr(request.user, 'usuario_sistema', None),
    })

@login_required
def venda_comunicacoes(request, venda_id):
    """Exibe as comunicações de uma venda específica"""
    if not (request.user.is_superuser or request.user.has_perm('core.view_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para visualizar comunicações de venda.'})
    
    venda = get_object_or_404(Venda, pk=venda_id)
    comunicacoes = venda.comunicacoes.all().order_by('-data_criacao')
    
    return render(request, 'core/venda_comunicacoes.html', {
        'venda': venda,
        'comunicacoes': comunicacoes,
        'usuario_sistema': request.user
    })

@login_required
def venda_criar_comunicacoes_obrigatorias(request, venda_id):
    """Cria as comunicações obrigatórias para uma venda"""
    if not (request.user.is_superuser or request.user.has_perm('core.add_comunicacaovenda')):
        return render(request, 'core/acesso_negado.html', {'mensagem': 'Você não tem permissão para criar comunicações de venda.'})
    
    venda = get_object_or_404(Venda, pk=venda_id)
    
    if request.method == 'POST':
        venda.criar_comunicacoes_obrigatorias(request.user.usuario_sistema)
        messages.success(request, 'Comunicações obrigatórias criadas com sucesso!')
        return redirect('core:venda_comunicacoes', venda_id=venda_id)
    
    return render(request, 'core/venda_criar_comunicacoes_obrigatorias.html', {
        'venda': venda,
        'usuario_sistema': request.user
    })

# ============================================================================
# VIEWS DE NOTIFICAÇÕES
# ============================================================================

@login_required
def notificacao_list(request):
    """Lista de notificações do usuário"""
    notificacoes = request.user.usuario_sistema.notificacoes.all()
    
    # Marcar como lidas se solicitado
    if request.GET.get('marcar_lidas'):
        notificacoes.filter(lida=False).update(lida=True)
        messages.success(request, 'Notificações marcadas como lidas!')
        return redirect('core:notificacao_list')
    
    return render(request, 'core/notificacao_list.html', {
        'notificacoes': notificacoes,
        'usuario_sistema': request.user
    })

@login_required
def notificacao_marcar_lida(request, pk):
    """Marca uma notificação como lida"""
    from .models import Notificacao
    notificacao = get_object_or_404(Notificacao, pk=pk, usuario=request.user.usuario_sistema)
    
    if request.method == 'POST':
        notificacao.lida = True
        notificacao.save()
        messages.success(request, 'Notificação marcada como lida!')
        
        # Redirecionar para o link da notificação se existir
        if notificacao.link:
            return redirect(notificacao.link)
        return redirect('core:notificacao_list')
    
    return render(request, 'core/notificacao_confirm_lida.html', {
        'notificacao': notificacao,
        'usuario_sistema': request.user
    })

@login_required
def notificacao_delete(request, pk):
    """Exclui uma notificação"""
    from .models import Notificacao
    notificacao = get_object_or_404(Notificacao, pk=pk, usuario=request.user.usuario_sistema)
    
    if request.method == 'POST':
        notificacao.delete()
        messages.success(request, 'Notificação excluída!')
        return redirect('core:notificacao_list')
    
    return render(request, 'core/notificacao_confirm_delete.html', {
        'notificacao': notificacao,
        'usuario_sistema': request.user
    })

@login_required
def notificacao_count(request):
    """Retorna o número de notificações não lidas (para AJAX)"""
    from .models import Notificacao
    count = Notificacao.objects.filter(
        usuario=request.user.usuario_sistema,
        lida=False
    ).count()
    
    return JsonResponse({'count': count})